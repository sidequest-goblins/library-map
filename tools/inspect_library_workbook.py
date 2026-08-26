import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


LIST_VIEW_WORKBOOK_PATH = Path(
    "C:/Users/cjade/OneDrive/"
    "Shared Workbooks/MyLibrary/"
    "LIBRARY LIST VIEW.xlsx"
)

REFERENCE_WORKBOOK_PATH = Path(
    "C:/Users/cjade/OneDrive/"
    "Shared Workbooks/MyLibrary/"
    "LIBRARY.xlsx"
)

LIST_VIEW_EXPECTED_HEADERS = (
    "CJ",
    "JC",
    "BIPOC",
    "LGBTQ+",
    "ISBN",
    "Year",
    "Pages",
    "Title",
    "Series",
    "First",
    "Last",
    "Genre",
    "Subgenre",
    "Publisher",
    "Origin",
    "Bookcase",
    "Shelf",
    "Position",
    "SYSTEM COLUMNS - AUTOMATION ONLY",
    "Needs Review",
    "Book ID",
    "Series Sort",
    "Volume Sort",
    "Last Sort",
    "First Sort",
)

BOOK_ID_PATTERN = re.compile(
    r"book-[0-9a-f]{8}-"
    r"[0-9a-f]{4}-"
    r"[0-9a-f]{4}-"
    r"[0-9a-f]{4}-"
    r"[0-9a-f]{12}"
)



def clean(value: Any) -> str:
    return str(value or "").strip()


def normalize_header(value: Any) -> str:
    return clean(value).lower().replace(" ", "")


def is_missing(value: Any) -> bool:
    normalized_value = clean(value).lower()

    return normalized_value in {
        "",
        "unknown",
        "n/a",
        "na",
        "none",
    }


def get_header_index(
    header_indexes: dict[str, int],
    *header_names: str,
) -> int | None:
    for header_name in header_names:
        index = header_indexes.get(
            normalize_header(header_name)
        )

        if index is not None:
            return index

    return None


def get_value(
    row_values: list[Any],
    header_indexes: dict[str, int],
    *header_names: str,
) -> Any:
    index = get_header_index(
        header_indexes,
        *header_names,
    )

    if (
        index is None
        or index >= len(row_values)
    ):
        return ""

    return row_values[index]


def validate_list_view_headers(
    sheet,
) -> None:
    header_row = next(
        sheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        ),
        (),
    )

    actual_headers = [
        clean(header)
        for header in header_row
    ]

    while (
        actual_headers
        and not actual_headers[-1]
    ):
        actual_headers.pop()

    expected_headers = list(
        LIST_VIEW_EXPECTED_HEADERS
    )

    if len(actual_headers) != len(
        expected_headers
    ):
        raise ValueError(
            f"The List View header count does not match. "
            f"Expected {len(expected_headers)} headers, "
            f"but found {len(actual_headers)}."
        )

    for column_number, (
        actual_header,
        expected_header,
    ) in enumerate(
        zip(
            actual_headers,
            expected_headers,
        ),
        start=1,
    ):
        if (
            normalize_header(actual_header)
            != normalize_header(
                expected_header
            )
        ):
            raise ValueError(
                "List View header mismatch at "
                f"column {column_number}. "
                f"Expected '{expected_header}', "
                f"but found '{actual_header}'."
            )


def find_list_view_sheet(workbook):
    """
    Prefer the explicitly named List View sheet.

    If the sheet is ever renamed or moved into another workbook,
    fall back to identifying it by its complete expected header
    structure.
    """
    if "List View" in workbook.sheetnames:
        sheet = workbook["List View"]

        validate_list_view_headers(
            sheet
        )

        return sheet

    for sheet in workbook.worksheets:
        try:
            validate_list_view_headers(
                sheet
            )
        except ValueError:
            continue

        return sheet

    raise ValueError(
        "Could not find a sheet with the complete "
        "expected List View header structure."
    )


def load_bookcase_rooms(
    workbook,
) -> dict[str, str]:
    if "Bookcases" not in workbook.sheetnames:
        raise ValueError(
            "Could not find the Bookcases sheet."
        )

    sheet = workbook["Bookcases"]
    rows = sheet.iter_rows(
        values_only=True
    )

    try:
        headers = next(rows)
    except StopIteration:
        raise ValueError(
            "The Bookcases sheet is empty."
        )

    header_indexes = {
        normalize_header(header): index
        for index, header in enumerate(
            headers
        )
    }

    bookcase_rooms: dict[str, str] = {}

    for row in rows:
        row_values = list(row)

        bookcase = clean(
            get_value(
                row_values,
                header_indexes,
                "bookcase",
            )
        )

        room = clean(
            get_value(
                row_values,
                header_indexes,
                "room",
            )
        )

        if bookcase:
            bookcase_rooms[bookcase] = room

    return bookcase_rooms


def format_count(
    missing_count: int,
    total_count: int,
) -> str:
    complete_count = (
        total_count - missing_count
    )

    if total_count <= 0:
        percentage = 0
    else:
        percentage = round(
            complete_count
            / total_count
            * 100,
            1,
        )

    return (
        f"{missing_count:>4} missing"
        f"  |  "
        f"{complete_count:>4} complete"
        f"  |  "
        f"{percentage:>5.1f}% complete"
    )


def main() -> None:
    for workbook_path in (
        LIST_VIEW_WORKBOOK_PATH,
        REFERENCE_WORKBOOK_PATH,
    ):
        if not workbook_path.exists():
            raise FileNotFoundError(
                f"Could not find workbook: "
                f"{workbook_path}"
            )

    print(
        "Loading List View workbook: "
        f"{LIST_VIEW_WORKBOOK_PATH}"
    )

    list_view_workbook = load_workbook(
        LIST_VIEW_WORKBOOK_PATH,
        read_only=True,
        data_only=True,
    )

    print(
        "Loading reference workbook: "
        f"{REFERENCE_WORKBOOK_PATH}"
    )

    reference_workbook = load_workbook(
        REFERENCE_WORKBOOK_PATH,
        read_only=True,
        data_only=True,
    )

    list_view_sheet = (
        find_list_view_sheet(
            list_view_workbook
        )
    )

    bookcase_rooms = (
        load_bookcase_rooms(
            list_view_workbook
        )
    )

    print(
        f"Using List View sheet: "
        f"{list_view_sheet.title}"
    )

    print(
        "Using Bookcases sheet from: "
        f"{LIST_VIEW_WORKBOOK_PATH.name}"
    )

    rows = list_view_sheet.iter_rows(
        values_only=True
    )

    try:
        headers = next(rows)
    except StopIteration:
        raise ValueError(
            "The List View sheet is empty."
        )

    header_indexes = {
        normalize_header(header): index
        for index, header in enumerate(
            headers
        )
    }

    year_header_index = (
        get_header_index(
            header_indexes,
            "year",
            "publication year",
            "publicationyear",
        )
    )

    isbn_header_index = (
        get_header_index(
            header_indexes,
            "isbn",
            "isbn-13",
            "isbn13",
        )
    )

    position_header_index = (
        get_header_index(
            header_indexes,
            "position",
            "shelf position",
            "shelfposition",
        )
    )

    boundary_header_index = (
        get_header_index(
            header_indexes,
            "SYSTEM COLUMNS - AUTOMATION ONLY",
        )
    )

    needs_review_header_index = (
        get_header_index(
            header_indexes,
            "Needs Review",
            "needsreview",
        )
    )

    book_id_header_index = (
        get_header_index(
            header_indexes,
            "Book ID",
            "bookid",
        )
    )

    if year_header_index is None:
        raise ValueError(
            "Could not find the Year column "
            "in List View."
        )

    if isbn_header_index is None:
        raise ValueError(
            "Could not find the ISBN column "
            "in List View."
        )

    if position_header_index is None:
        raise ValueError(
            "Could not find the Position column "
            "in List View."
        )

    if boundary_header_index is None:
        raise ValueError(
            "Could not find the exact system "
            "boundary column in List View."
        )

    if needs_review_header_index is None:
        raise ValueError(
            "Could not find the Needs Review column "
            "in List View."
        )

    if book_id_header_index is None:
        raise ValueError(
            "Could not find the Book ID column "
            "in List View."
        )

    if (
        needs_review_header_index
        != boundary_header_index + 1
    ):
        raise ValueError(
            "Needs Review must appear immediately after "
            "SYSTEM COLUMNS - AUTOMATION ONLY."
        )

    if (
        book_id_header_index
        != needs_review_header_index + 1
    ):
        raise ValueError(
            "Book ID must appear immediately after "
            "Needs Review."
        )


    total_books = 0

    missing_year = 0
    missing_isbn = 0

    missing_book_id = 0
    malformed_book_id = 0
    duplicate_book_id = 0
    orphan_book_id = 0

    missing_bookcase = 0
    missing_shelf = 0
    missing_room = 0
    incomplete_location = 0

    office_books = 0
    missing_office_position = 0

    fully_complete = 0

    incomplete_rows: list[
        dict[str, Any]
    ] = []

    book_id_issues: list[
        dict[str, Any]
    ] = []

    book_id_first_rows: dict[
        str,
        int,
    ] = {}


    for row_number, row in enumerate(
        rows,
        start=2,
    ):
        row_values = list(row)

        title = clean(
            get_value(
                row_values,
                header_indexes,
                "title",
            )
        )

        book_id = clean(
            get_value(
                row_values,
                header_indexes,
                "Book ID",
                "bookid",
            )
        )

        if not title:
            if book_id:
                orphan_book_id += 1

                book_id_issues.append(
                    {
                        "row": row_number,
                        "title": "(blank title)",
                        "issue": (
                            "Book ID exists on a "
                            "blank Title row"
                        ),
                    }
                )

            continue

        total_books += 1

        year = get_value(
            row_values,
            header_indexes,
            "year",
            "publication year",
            "publicationyear",
        )

        isbn = get_value(
            row_values,
            header_indexes,
            "isbn",
            "isbn-13",
            "isbn13",
        )

        bookcase = clean(
            get_value(
                row_values,
                header_indexes,
                "bookcase",
            )
        )

        shelf = clean(
            get_value(
                row_values,
                header_indexes,
                "shelf",
            )
        )

        position = get_value(
            row_values,
            header_indexes,
            "position",
            "shelf position",
            "shelfposition",
        )

        room = (
            bookcase_rooms.get(
                bookcase,
                "",
            )
            if bookcase
            else ""
        )

        row_missing_year = (
            is_missing(year)
        )

        row_missing_isbn = (
            is_missing(isbn)
        )

        row_missing_book_id = (
            is_missing(book_id)
        )

        if row_missing_book_id:
            missing_book_id += 1

            book_id_issues.append(
                {
                    "row": row_number,
                    "title": title,
                    "issue": "Missing Book ID",
                }
            )
        else:
            if (
                BOOK_ID_PATTERN.fullmatch(
                    book_id
                )
                is None
            ):
                malformed_book_id += 1

                book_id_issues.append(
                    {
                        "row": row_number,
                        "title": title,
                        "issue": (
                            "Malformed Book ID: "
                            f"{book_id}"
                        ),
                    }
                )

            normalized_book_id = (
                book_id.lower()
            )

            first_row = (
                book_id_first_rows.get(
                    normalized_book_id
                )
            )

            if first_row is None:
                book_id_first_rows[
                    normalized_book_id
                ] = row_number
            else:
                duplicate_book_id += 1

                book_id_issues.append(
                    {
                        "row": row_number,
                        "title": title,
                        "issue": (
                            "Duplicate Book ID; "
                            f"first used on row "
                            f"{first_row}"
                        ),
                    }
                )

        row_missing_bookcase = (
            is_missing(bookcase)
        )

        row_missing_shelf = (
            is_missing(shelf)
        )

        row_missing_room = (
            is_missing(room)
        )

        row_incomplete_location = (
            row_missing_room
            or row_missing_bookcase
            or row_missing_shelf
        )

        is_office_book = (
            bookcase.lower()
            == "office"
        )

        row_missing_office_position = (
            is_office_book
            and is_missing(position)
        )

        if row_missing_year:
            missing_year += 1

        if row_missing_isbn:
            missing_isbn += 1

        if row_missing_bookcase:
            missing_bookcase += 1

        if row_missing_shelf:
            missing_shelf += 1

        if row_missing_room:
            missing_room += 1

        if row_incomplete_location:
            incomplete_location += 1

        if is_office_book:
            office_books += 1

        if row_missing_office_position:
            missing_office_position += 1

        row_is_fully_complete = (
            not row_missing_year
            and not row_missing_isbn
            and not row_incomplete_location
            and not row_missing_office_position
        )

        if row_is_fully_complete:
            fully_complete += 1
        else:
            missing_fields: list[str] = []

            if row_missing_year:
                missing_fields.append(
                    "Year"
                )

            if row_missing_isbn:
                missing_fields.append(
                    "ISBN"
                )

            if row_missing_room:
                missing_fields.append(
                    "Room"
                )

            if row_missing_bookcase:
                missing_fields.append(
                    "Bookcase"
                )

            if row_missing_shelf:
                missing_fields.append(
                    "Shelf"
                )

            if row_missing_office_position:
                missing_fields.append(
                    "Office position"
                )

            incomplete_rows.append(
                {
                    "row": row_number,
                    "title": title,
                    "missing": missing_fields,
                }
            )

    print()
    print("=" * 68)
    print("LIBRARY WORKBOOK COVERAGE")
    print("=" * 68)

    print(
        f"\nTotal List View books: "
        f"{total_books}"
    )

    book_id_integrity_passed = (
        missing_book_id == 0
        and malformed_book_id == 0
        and duplicate_book_id == 0
        and orphan_book_id == 0
    )

    print("\nSystem integrity:")
    print(
        "  Book ID assigned:  "
        + format_count(
            missing_book_id,
            total_books,
        )
    )

    print(
        f"  Malformed IDs:     "
        f"{malformed_book_id}"
    )

    print(
        f"  Duplicate IDs:     "
        f"{duplicate_book_id}"
    )

    print(
        f"  IDs on blank rows: "
        f"{orphan_book_id}"
    )

    print(
        "  Integrity status:  "
        + (
            "PASS"
            if book_id_integrity_passed
            else "FAIL"
        )
    )

    print("\nCore book data:")
    print(
        "  Year:              "
        + format_count(
            missing_year,
            total_books,
        )
    )

    print(
        "  ISBN:              "
        + format_count(
            missing_isbn,
            total_books,
        )
    )

    print("\nLocation data:")
    print(
        "  Complete location: "
        + format_count(
            incomplete_location,
            total_books,
        )
    )

    print(
        "  Room:              "
        + format_count(
            missing_room,
            total_books,
        )
    )

    print(
        "  Bookcase:          "
        + format_count(
            missing_bookcase,
            total_books,
        )
    )

    print(
        "  Shelf:             "
        + format_count(
            missing_shelf,
            total_books,
        )
    )

    print("\nOffice-only data:")
    print(
        f"  Office books:      "
        f"{office_books}"
    )

    print(
        "  Position:          "
        + format_count(
            missing_office_position,
            office_books,
        )
    )

    fully_complete_missing = (
        total_books
        - fully_complete
    )

    print("\nOverall tracked coverage:")
    print(
        "  All fields:        "
        + format_count(
            fully_complete_missing,
            total_books,
        )
    )

    if book_id_issues:
        print(
            "\nFirst 20 Book ID integrity problems:"
        )

        for item in book_id_issues[:20]:
            print(
                f"  Row {item['row']}: "
                f"{item['title']} "
                f"- {item['issue']}"
            )

        remaining_book_id_issues = (
            len(book_id_issues)
            - 20
        )

        if remaining_book_id_issues > 0:
            print(
                f"  ...and "
                f"{remaining_book_id_issues} more"
            )

    if incomplete_rows:
        print(
            "\nFirst 20 books needing work:"
        )

        for item in incomplete_rows[:20]:
            missing_label = ", ".join(
                item["missing"]
            )

            print(
                f"  Row {item['row']}: "
                f"{item['title']} "
                f"— {missing_label}"
            )

        remaining_count = (
            len(incomplete_rows)
            - 20
        )

        if remaining_count > 0:
            print(
                f"  ...and "
                f"{remaining_count} more"
            )
    else:
        print(
            "\n✨ Every tracked field is complete."
        )

    print()
    print("=" * 68)

    list_view_workbook.close()
    reference_workbook.close()


if __name__ == "__main__":
    main()
