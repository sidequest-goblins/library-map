import hashlib
import json
import posixpath
import re
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from PIL import Image, ImageFile, UnidentifiedImageError

ImageFile.LOAD_TRUNCATED_IMAGES = True

CATALOG_WORKBOOK_PATH = Path(
    "C:/Users/cjade/OneDrive/Shared Workbooks/MyLibrary/LIBRARY.xlsx"
)

LIST_VIEW_WORKBOOK_PATH = Path(
    "C:/Users/cjade/OneDrive/Shared Workbooks/MyLibrary/LIBRARY LIST VIEW.xlsx"
)
OUTPUT_DIR = Path("C:/library_app/library-map/public/data")
BOOKS_OUTPUT_PATH = OUTPUT_DIR / "library-books.json"
CATALOG_OUTPUT_PATH = OUTPUT_DIR / "library-catalog.json"
WANTED_OUTPUT_PATH = OUTPUT_DIR / "library-wanted.json"
META_OUTPUT_PATH = OUTPUT_DIR / "library-meta.json"
CATALOG_REQUIRED_HEADERS = {
    "cover",
    "title",
    "first",
    "last",
    "format",
}
DEBUG_IMAGE_INSPECTION = False
COVER_OUTPUT_DIR = OUTPUT_DIR / "covers"
COVER_PUBLIC_PATH = "/data/covers"
COVER_IMAGE_EXTENSION = "webp"
COVER_WEBP_QUALITY = 76
COVER_CACHE_PATH = OUTPUT_DIR / "library-cover-cache.json"
COVER_CACHE_VERSION = 1

def clean(value: Any) -> str:
    return str(value or "").strip()

def split_name_parts(value: str) -> list[str]:
    text = str(value or "")
    return [clean(part) for part in re.split(r"[;\r\n]+", text) if clean(part)]

def make_author(first: str, last: str) -> str:
    first_parts = split_name_parts(first)
    last_parts = split_name_parts(last)

    if not first_parts and not last_parts:
        return ""

    max_count = max(len(first_parts), len(last_parts))

    authors: list[str] = []

    for index in range(max_count):
        first_part = first_parts[index] if index < len(first_parts) else ""
        last_part = last_parts[index] if index < len(last_parts) else ""

        if first_part and last_part:
            authors.append(f"{first_part} {last_part}")
        else:
            authors.append(first_part or last_part)

    return "; ".join(authors)

def make_author_sort(first: str, last: str) -> str:
    first_parts = split_name_parts(first)
    last_parts = split_name_parts(last)

    first = clean(first)
    last = clean(last)

    # Keep alphabetical sorting based on the first listed last name.
    if first_parts and last_parts:
        return f"{last_parts[0]}, {first_parts[0]}"

    if last_parts:
        return last_parts[0]

    if first_parts:
        return first_parts[0]

    return last or first

def checkbox_to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value

    text = clean(value).lower()
    return text in {"true", "yes", "y", "1", "checked", "x"}

def load_bookcase_rooms(workbook) -> dict[str, str]:
    if "Bookcases" not in workbook.sheetnames:
        raise ValueError("Could not find required sheet: Bookcases")

    sheet = workbook["Bookcases"]
    rows = sheet.iter_rows(values_only=True)

    try:
        headers = [clean(value) for value in next(rows)]
    except StopIteration:
        raise ValueError("The Bookcases sheet is empty.")

    header_indexes = {normalize_header(header): index for index, header in enumerate(headers)}

    required_headers = ["bookcase", "room"]
    missing_headers = [
        header for header in required_headers if header not in header_indexes
    ]

    if missing_headers:
        raise ValueError(
            "Bookcases sheet is missing expected normalized headers: "
            + ", ".join(missing_headers)
            + f"\nFound headers: {headers}"
        )

    bookcase_rooms = {}

    for row in rows:
        row_values = list(row)

        def get(header: str) -> str:
            index = header_indexes[header]
            if index >= len(row_values):
                return ""
            return clean(row_values[index])

        bookcase = get("bookcase")
        room = get("room")

        if bookcase and room:
            bookcase_rooms[bookcase] = room

    return bookcase_rooms

def get_room_for_bookcase(bookcase: str, bookcase_rooms: dict[str, str]) -> str:
    bookcase = clean(bookcase)
    return bookcase_rooms.get(bookcase, "")

def parse_optional_int(value: Any) -> int | None:
    text = clean(value)

    if not text:
        return None

    try:
        return int(float(text))
    except ValueError:
        return None

SeriesNumber = int | float

def parse_series_number(value: str) -> SeriesNumber:
    number = float(value)
    return int(number) if number.is_integer() else number

def parse_optional_series_number(value: Any) -> SeriesNumber | None:
    text = clean(value)

    if not text:
        return None

    number_match = re.search(r"\d+(?:\.\d+)?", text)

    if not number_match:
        return None

    return parse_series_number(number_match.group(0))

def parse_series_cell(value: Any) -> dict[str, str | SeriesNumber | None]:
    text = clean(value)

    if not text:
        return {
            "series": None,
            "seriesTitle": None,
            "seriesNumber": None,
        }

    # Supports List View values like:
    # "True Colors #1"
    # "The Black Jewels #1-3"
    # "The Roots of Chaos #0.1"
    series_match = re.match(
        r"^(.*?)\s*#\s*(\d+(?:\.\d+)?)(?:\s*-\s*\d+(?:\.\d+)?)?\s*$",
        text,
    )

    if series_match:
        series_title = series_match.group(1).strip()
        series_number = parse_series_number(series_match.group(2))

        return {
            "series": series_title or None,
            "seriesTitle": series_title or None,
            "seriesNumber": series_number,
        }

    return {
        "series": text,
        "seriesTitle": text,
        "seriesNumber": None,
    }
    
def parse_shelf(raw_shelf: str) -> tuple[str, str]:
    raw_shelf = clean(raw_shelf)

    if not raw_shelf:
        return "", "Main"

    parts = [part.strip() for part in raw_shelf.split(",") if part.strip()]

    shelf = parts[0] if parts else raw_shelf
    row = parts[1] if len(parts) > 1 else "Main"

    return shelf, row

def parse_title(raw_title: str) -> dict[str, str | SeriesNumber | None]:
    title = clean(raw_title)

    # Manga:
    # "Rurouni Kenshin, Vol. 2"
    # "Tokyo Ghoul: re, Vol. 14"
    # Also supports decimals/ranges like:
    # "Series, Vol. 0.5"
    # "Series, Vol. 12.5-14.15"
    volume_match = re.match(
        r"^(.*?),\s*Vol\.?\s*(\d+(?:\.\d+)?)(?:\s*-\s*\d+(?:\.\d+)?)?(?:\s*\((Light Novel|Manwha|Manhwa|Manga)\))?\s*$",
        title,
        re.IGNORECASE,
    )
    if volume_match:
        series = volume_match.group(1).strip()
        series_number = parse_series_number(volume_match.group(2))
        format_label = volume_match.group(3).strip() if volume_match.group(3) else None

        series_key = f"{series}|{format_label}" if format_label else series

        return {
            "title": title,
            "rawTitle": title,
            "series": series_key,
            "seriesTitle": series,
            "seriesFormat": format_label,
            "seriesNumber": series_number,
        }

    # Regular books:
    # "The Hunger Games (The Hunger Games #1)"
    # "Queen B (HMRC #0.5)"
    # "Goblin Crimes (Some Series #12.5-14.15)"
    #
    # For ranges, seriesNumber uses the first number.
    # Example: #12.5-14.15 -> 12.5
    book_match = re.match(
        r"^(.*?)\s*\((.*?)\s*#\s*(\d+(?:\.\d+)?)(?:\s*-\s*\d+(?:\.\d+)?)?\)\s*$",
        title,
    )
    if book_match:
        clean_title = book_match.group(1).strip()
        series = book_match.group(2).strip()
        series_number = parse_series_number(book_match.group(3))

        return {
            "title": clean_title or title,
            "rawTitle": title,
            "series": series or None,
            "seriesTitle": series or None,
            "seriesFormat": None,
            "seriesNumber": series_number,
        }

    return {
        "title": title,
        "rawTitle": title,
        "series": None,
        "seriesTitle": None,
        "seriesFormat": None,
        "seriesNumber": None,
    }

def normalize_header(value: Any) -> str:
    return clean(value).lower().replace(" ", "")

def get_header_index(header_indexes: dict[str, int], *headers: str) -> int | None:
    for header in headers:
        index = header_indexes.get(normalize_header(header))

        if index is not None:
            return index

    return None

def get_value_by_header_alias(
    row_values: list[Any],
    header_indexes: dict[str, int],
    *headers: str,
) -> Any:
    index = get_header_index(header_indexes, *headers)

    if index is None or index >= len(row_values):
        return ""

    return row_values[index]

def find_sheet_with_headers(
    workbook,
    required_headers: set[str],
    label: str,
):
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)

        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration:
            continue

        normalized_headers = {normalize_header(header) for header in headers}

        if required_headers.issubset(normalized_headers):
            return sheet

    raise ValueError(f"Could not find {label} sheet with headers: {required_headers}")

def get_sheet_headers(sheet) -> list[str]:
    rows = sheet.iter_rows(values_only=True)

    try:
        return [clean(value) for value in next(rows)]
    except StopIteration:
        return []

def is_catalog_sheet(sheet) -> bool:
    headers = get_sheet_headers(sheet)
    normalized_headers = {normalize_header(header) for header in headers}

    return CATALOG_REQUIRED_HEADERS.issubset(normalized_headers)

def get_image_anchor_cell(image) -> tuple[int | None, int | None]:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)

    if marker is None:
        return None, None

    # openpyxl stores these as zero-based indexes.
    return marker.row + 1, marker.col + 1

def get_image_bytes(image) -> bytes:
    data = image._data

    if callable(data):
        return data()

    return data

def guess_image_extension(image, image_bytes: bytes) -> str:
    image_format = clean(getattr(image, "format", "")).lower()

    if image_format in {"jpeg", "jpg"}:
        return "jpg"

    if image_format == "png":
        return "png"

    if image_bytes.startswith(b"\x89PNG"):
        return "png"

    if image_bytes.startswith(b"\xff\xd8"):
        return "jpg"

    if image_bytes.startswith(b"GIF8"):
        return "gif"

    return image_format or "unknown"

def inspect_catalog_images(workbook) -> dict[str, Any]:
    print("\nInspecting embedded catalog images:")

    report: dict[str, Any] = {
        "totalImages": 0,
        "totalImageBytes": 0,
        "imagesBySheet": {},
        "formats": {},
        "missingAnchors": [],
    }

    formats = Counter()

    for sheet in workbook.worksheets:
        if not is_catalog_sheet(sheet):
            continue

        images = list(getattr(sheet, "_images", []))
        sheet_image_bytes = 0
        anchored_rows: list[int] = []

        print(f"\n  {sheet.title}: {len(images)} image(s)")

        for image_index, image in enumerate(images, start=1):
            row_number, col_number = get_image_anchor_cell(image)
            image_bytes = get_image_bytes(image)
            image_size = len(image_bytes)
            extension = guess_image_extension(image, image_bytes)

            formats[extension] += 1
            sheet_image_bytes += image_size
            report["totalImages"] += 1
            report["totalImageBytes"] += image_size

            if row_number is not None:
                anchored_rows.append(row_number)
            else:
                report["missingAnchors"].append(
                    {
                        "sheet": sheet.title,
                        "imageIndex": image_index,
                    }
                )

            if image_index <= 8:
                size_kb = image_size / 1024
                print(
                    f"    image {image_index}: "
                    f"row={row_number}, col={col_number}, "
                    f"format={extension}, size={size_kb:.1f} KB"
                )

        if len(images) > 8:
            print(f"    ... {len(images) - 8} more image(s)")

        duplicate_anchor_rows = sorted(
            row for row, count in Counter(anchored_rows).items() if count > 1
        )

        report["imagesBySheet"][sheet.title] = {
            "imageCount": len(images),
            "imageBytes": sheet_image_bytes,
            "firstAnchoredRows": anchored_rows[:12],
            "duplicateAnchorRows": duplicate_anchor_rows,
            "missingAnchorCount": len(images) - len(anchored_rows),
        }

        print(f"    first rows: {anchored_rows[:12]}")
        if duplicate_anchor_rows:
            print(f"    duplicate anchor rows: {duplicate_anchor_rows}")

    report["formats"] = dict(sorted(formats.items()))

    total_mb = report["totalImageBytes"] / 1024 / 1024
    print("\nEmbedded image summary:")
    print(f"  total images: {report['totalImages']}")
    print(f"  total embedded image size: {total_mb:.2f} MB")
    print(f"  formats: {report['formats']}")

    if report["missingAnchors"]:
        print(f"  missing anchors: {len(report['missingAnchors'])}")

    return report

def inspect_workbook_image_package(workbook_path: Path) -> dict[str, Any]:
    print("\nInspecting workbook package for stored images:")

    report: dict[str, Any] = {
        "mediaFileCount": 0,
        "mediaTotalBytes": 0,
        "mediaExtensions": {},
        "drawingFileCount": 0,
        "cellImageRelatedFiles": [],
        "externalLinkRelatedFiles": [],
    }

    with zipfile.ZipFile(workbook_path, "r") as archive:
        names = archive.namelist()

        media_files = [
            name for name in names
            if name.startswith("xl/media/")
        ]

        drawing_files = [
            name for name in names
            if name.startswith("xl/drawings/") and name.endswith(".xml")
        ]

        cell_image_related_files = [
            name for name in names
            if "cellimage" in name.lower()
            or "richdata" in name.lower()
            or "richvalue" in name.lower()
        ]

        external_link_related_files = [
            name for name in names
            if "external" in name.lower()
            or "externalLink" in name
        ]

        extensions = Counter()
        total_bytes = 0

        for name in media_files:
            extension = Path(name).suffix.lower().lstrip(".") or "unknown"
            extensions[extension] += 1
            total_bytes += archive.getinfo(name).file_size

        report["mediaFileCount"] = len(media_files)
        report["mediaTotalBytes"] = total_bytes
        report["mediaExtensions"] = dict(sorted(extensions.items()))
        report["drawingFileCount"] = len(drawing_files)
        report["cellImageRelatedFiles"] = cell_image_related_files
        report["externalLinkRelatedFiles"] = external_link_related_files

        print(f"  xl/media files: {len(media_files)}")
        print(f"  xl/media total size: {total_bytes / 1024 / 1024:.2f} MB")
        print(f"  media extensions: {report['mediaExtensions']}")
        print(f"  drawing xml files: {len(drawing_files)}")

        if media_files:
            print("  first media files:")
            for name in media_files[:12]:
                size_kb = archive.getinfo(name).file_size / 1024
                print(f"    {name} ({size_kb:.1f} KB)")

            if len(media_files) > 12:
                print(f"    ... {len(media_files) - 12} more")

        if cell_image_related_files:
            print("  possible in-cell image / rich data files:")
            for name in cell_image_related_files[:20]:
                print(f"    {name}")

            if len(cell_image_related_files) > 20:
                print(f"    ... {len(cell_image_related_files) - 20} more")

        if external_link_related_files:
            print("  possible external-link files:")
            for name in external_link_related_files[:20]:
                print(f"    {name}")

            if len(external_link_related_files) > 20:
                print(f"    ... {len(external_link_related_files) - 20} more")

    return report

def normalize_xlsx_path(base_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")

    base_dir = posixpath.dirname(base_path)
    return posixpath.normpath(posixpath.join(base_dir, target))

def read_xml(archive: zipfile.ZipFile, path: str):
    return ET.fromstring(archive.read(path))

def get_relationships(archive: zipfile.ZipFile, source_path: str) -> dict[str, str]:
    source_dir = posixpath.dirname(source_path)
    source_name = posixpath.basename(source_path)
    rels_path = posixpath.join(source_dir, "_rels", f"{source_name}.rels")

    if rels_path not in archive.namelist():
        return {}

    root = read_xml(archive, rels_path)

    relationships: dict[str, str] = {}

    for rel in root:
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")

        if not rel_id or not target:
            continue

        relationships[rel_id] = normalize_xlsx_path(source_path, target)

    return relationships

def get_workbook_sheet_paths(archive: zipfile.ZipFile) -> dict[str, str]:
    namespace = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    workbook_path = "xl/workbook.xml"
    workbook_root = read_xml(archive, workbook_path)
    workbook_rels = get_relationships(archive, workbook_path)

    sheet_paths: dict[str, str] = {}

    for sheet in workbook_root.findall(".//main:sheet", namespace):
        sheet_name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(f"{{{namespace['rel']}}}id")

        if not sheet_name or not rel_id:
            continue

        sheet_path = workbook_rels.get(rel_id)

        if sheet_path:
            sheet_paths[sheet_name] = sheet_path

    return sheet_paths

def build_catalog_rows_by_sheet(
    catalog_books: dict[str, dict[str, Any]],
) -> dict[str, dict[int, dict[str, Any]]]:
    catalog_rows_by_sheet: dict[str, dict[int, dict[str, Any]]] = {}

    for catalog_book in catalog_books.values():
        source_sheet = catalog_book.get("sourceSheet")
        source_row = catalog_book.get("sourceRow")

        if not source_sheet or source_row is None:
            continue

        catalog_rows_by_sheet.setdefault(str(source_sheet), {})[int(source_row)] = catalog_book

    return catalog_rows_by_sheet

def inspect_drawing_image_anchors(
    workbook_path: Path,
    catalog_sheet_names: set[str],
    catalog_books: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    print("\nInspecting drawing anchors for catalog images:")

    namespace = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    report: dict[str, Any] = {
        "totalAnchoredImages": 0,
        "anchorsBySheet": {},
        "missingDrawingSheets": [],
        "missingMediaTargets": [],
        "catalogRowsWithoutAnchoredImages": [],
        "anchorRowsWithoutCatalogBooks": [],
        "duplicateMediaPaths": [],
    }

    catalog_rows_by_sheet = build_catalog_rows_by_sheet(catalog_books)

    media_usage: dict[str, list[dict[str, Any]]] = {}

    with zipfile.ZipFile(workbook_path, "r") as archive:
        sheet_paths = get_workbook_sheet_paths(archive)

        for sheet_name in sorted(catalog_sheet_names):
            sheet_path = sheet_paths.get(sheet_name)

            if not sheet_path:
                report["missingDrawingSheets"].append(sheet_name)
                print(f"\n  {sheet_name}: no worksheet path found")
                continue

            sheet_rels = get_relationships(archive, sheet_path)
            drawing_paths = [
                target
                for target in sheet_rels.values()
                if target.startswith("xl/drawings/") and target.endswith(".xml")
            ]

            if not drawing_paths:
                report["missingDrawingSheets"].append(sheet_name)
                print(f"\n  {sheet_name}: no drawing file linked")
                continue

            sheet_anchors: list[dict[str, Any]] = []

            for drawing_path in drawing_paths:
                drawing_root = read_xml(archive, drawing_path)
                drawing_rels = get_relationships(archive, drawing_path)

                anchor_nodes = (
                    drawing_root.findall(".//xdr:twoCellAnchor", namespace)
                    + drawing_root.findall(".//xdr:oneCellAnchor", namespace)
                    + drawing_root.findall(".//xdr:absoluteAnchor", namespace)
                )

                for anchor in anchor_nodes:
                    from_node = anchor.find("xdr:from", namespace)

                    row_number = None
                    col_number = None

                    if from_node is not None:
                        row_node = from_node.find("xdr:row", namespace)
                        col_node = from_node.find("xdr:col", namespace)

                        if row_node is not None and row_node.text is not None:
                            row_number = int(row_node.text) + 1

                        if col_node is not None and col_node.text is not None:
                            col_number = int(col_node.text) + 1

                    blip = anchor.find(".//a:blip", namespace)
                    rel_id = None

                    if blip is not None:
                        rel_id = blip.attrib.get(f"{{{namespace['rel']}}}embed")

                    media_path = drawing_rels.get(rel_id or "")

                    if not media_path:
                        report["missingMediaTargets"].append(
                            {
                                "sheet": sheet_name,
                                "drawing": drawing_path,
                                "row": row_number,
                                "col": col_number,
                                "relId": rel_id,
                            }
                        )

                    sheet_anchors.append(
                        {
                            "row": row_number,
                            "col": col_number,
                            "mediaPath": media_path,
                            "drawingPath": drawing_path,
                        }
                    )

                    if media_path:
                        media_usage.setdefault(media_path, []).append(
                            {
                                "sheet": sheet_name,
                                "row": row_number,
                                "col": col_number,
                            }
                        )

            sheet_anchors.sort(
                key=lambda item: (
                    item["row"] if item["row"] is not None else 999999,
                    item["col"] if item["col"] is not None else 999999,
                    item["mediaPath"] or "",
                )
            )

            duplicate_rows = sorted(
                row
                for row, count in Counter(
                    item["row"] for item in sheet_anchors if item["row"] is not None
                ).items()
                if count > 1
            )

            anchored_rows = {
                item["row"]
                for item in sheet_anchors
                if item["row"] is not None
            }

            catalog_rows = catalog_rows_by_sheet.get(sheet_name, {})

            missing_cover_rows = sorted(
                row_number
                for row_number in catalog_rows
                if row_number not in anchored_rows
            )

            extra_anchor_rows = sorted(
                row_number
                for row_number in anchored_rows
                if row_number not in catalog_rows
            )

            for row_number in missing_cover_rows:
                catalog_book = catalog_rows[row_number]
                report["catalogRowsWithoutAnchoredImages"].append(
                    {
                        "sheet": sheet_name,
                        "row": row_number,
                        "catalogKey": catalog_book["catalogKey"],
                        "title": catalog_book["title"],
                        "author": catalog_book["author"],
                    }
                )

            for row_number in extra_anchor_rows:
                report["anchorRowsWithoutCatalogBooks"].append(
                    {
                        "sheet": sheet_name,
                        "row": row_number,
                    }
                )

            report["anchorsBySheet"][sheet_name] = {
                "anchorCount": len(sheet_anchors),
                "firstAnchors": sheet_anchors[:12],
                "duplicateRows": duplicate_rows,
            }

            report["totalAnchoredImages"] += len(sheet_anchors)

            print(f"\n  {sheet_name}: {len(sheet_anchors)} anchored image(s)")

            for anchor in sheet_anchors[:8]:
                print(
                    f"    row={anchor['row']}, col={anchor['col']}, "
                    f"media={anchor['mediaPath']}"
                )

            if len(sheet_anchors) > 8:
                print(f"    ... {len(sheet_anchors) - 8} more")

            if duplicate_rows:
                print(f"    duplicate rows: {duplicate_rows}")
    duplicate_media_paths = {
        media_path: usages
        for media_path, usages in media_usage.items()
        if len(usages) > 1
    }

    for media_path, usages in sorted(duplicate_media_paths.items()):
        report["duplicateMediaPaths"].append(
            {
                "mediaPath": media_path,
                "usages": usages,
            }
        )

    print("\nDrawing anchor summary:")
    print(f"  total anchored images: {report['totalAnchoredImages']}")

    print(f"  unique anchored media files: {len(media_usage)}")

    if report["duplicateMediaPaths"]:
        print(f"  duplicate/reused media files: {len(report['duplicateMediaPaths'])}")

        for duplicate in report["duplicateMediaPaths"][:10]:
            print(f"    {duplicate['mediaPath']}")
            for usage in duplicate["usages"]:
                print(
                    f"      {usage['sheet']} row {usage['row']}, col {usage['col']}"
                )

        if len(report["duplicateMediaPaths"]) > 10:
            print(f"    ... {len(report['duplicateMediaPaths']) - 10} more")

    if report["missingDrawingSheets"]:
        print(f"  sheets without drawings: {report['missingDrawingSheets']}")

    if report["missingMediaTargets"]:
        print(f"  missing media targets: {len(report['missingMediaTargets'])}")

    return report

def get_cell_reference_parts(cell_reference: str) -> tuple[int | None, int | None]:
    match = re.match(r"^([A-Z]+)(\d+)$", cell_reference)

    if not match:
        return None, None

    column_letters = match.group(1)
    row_number = int(match.group(2))
    column_number = 0

    for letter in column_letters:
        column_number = column_number * 26 + (ord(letter) - ord("A") + 1)

    return row_number, column_number

def collect_catalog_cell_image_media_paths(
    workbook_path: Path,
    catalog_sheet_names: set[str],
) -> tuple[dict[tuple[str, int], str], dict[str, Any]]:
    namespace = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "richdata": "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata",
        "richvaluerel": "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    image_media_by_row: dict[tuple[str, int], str] = {}
    report: dict[str, Any] = {
        "totalCellImages": 0,
        "mappedCellImages": 0,
        "duplicateCellRows": [],
        "missingRichDataFiles": [],
        "missingRichValueIndexes": [],
        "missingRichRelationshipIndexes": [],
        "missingMediaTargets": [],
    }

    rich_value_rel_path = "xl/richData/richValueRel.xml"
    rich_value_path = "xl/richData/rdrichvalue.xml"

    with zipfile.ZipFile(workbook_path, "r") as archive:
        archive_names = set(archive.namelist())

        missing_rich_files = [
            path
            for path in [rich_value_rel_path, rich_value_path]
            if path not in archive_names
        ]

        if missing_rich_files:
            report["missingRichDataFiles"] = missing_rich_files
            return image_media_by_row, report

        rich_value_rel_root = read_xml(archive, rich_value_rel_path)
        rich_rel_ids = [
            rel.attrib.get(f"{{{namespace['rel']}}}id")
            for rel in rich_value_rel_root.findall("richvaluerel:rel", namespace)
        ]

        rich_relationships = get_relationships(archive, rich_value_rel_path)

        rich_value_root = read_xml(archive, rich_value_path)
        rich_value_media_paths: list[str | None] = []

        for rich_value_index, rich_value in enumerate(
            rich_value_root.findall("richdata:rv", namespace),
            start=1,
        ):
            value_nodes = rich_value.findall("richdata:v", namespace)

            if not value_nodes or value_nodes[0].text is None:
                report["missingRichRelationshipIndexes"].append(rich_value_index)
                rich_value_media_paths.append(None)
                continue

            rel_index = int(value_nodes[0].text)

            if rel_index >= len(rich_rel_ids):
                report["missingRichRelationshipIndexes"].append(rich_value_index)
                rich_value_media_paths.append(None)
                continue

            rel_id = rich_rel_ids[rel_index]
            media_path = rich_relationships.get(rel_id or "")

            if not media_path:
                report["missingMediaTargets"].append(
                    {
                        "richValueIndex": rich_value_index,
                        "relIndex": rel_index,
                        "relId": rel_id,
                    }
                )

            rich_value_media_paths.append(media_path)

        sheet_paths = get_workbook_sheet_paths(archive)

        for sheet_name in sorted(catalog_sheet_names):
            sheet_path = sheet_paths.get(sheet_name)

            if not sheet_path:
                continue

            sheet_root = read_xml(archive, sheet_path)

            for cell in sheet_root.findall(".//main:c", namespace):
                vm_index_text = cell.attrib.get("vm")
                cell_reference = cell.attrib.get("r")

                if not vm_index_text or not cell_reference:
                    continue

                row_number, column_number = get_cell_reference_parts(cell_reference)

                # Catalog cover images live in column A. Ignore any future rich-data
                # cells elsewhere on the sheet so we do not attach the wrong media.
                if row_number is None or column_number != 1:
                    continue

                report["totalCellImages"] += 1

                rich_value_index = int(vm_index_text)
                media_index = rich_value_index - 1

                if media_index < 0 or media_index >= len(rich_value_media_paths):
                    report["missingRichValueIndexes"].append(
                        {
                            "sheet": sheet_name,
                            "row": row_number,
                            "cell": cell_reference,
                            "richValueIndex": rich_value_index,
                        }
                    )
                    continue

                media_path = rich_value_media_paths[media_index]

                if not media_path:
                    continue

                key = (sheet_name, row_number)

                if key in image_media_by_row:
                    report["duplicateCellRows"].append(
                        {
                            "sheet": sheet_name,
                            "row": row_number,
                            "firstMediaPath": image_media_by_row[key],
                            "secondMediaPath": media_path,
                        }
                    )
                    continue

                image_media_by_row[key] = media_path
                report["mappedCellImages"] += 1

    return image_media_by_row, report

def collect_catalog_image_media_paths(
    workbook_path: Path,
    catalog_sheet_names: set[str],
) -> tuple[dict[tuple[str, int], str], dict[str, Any]]:
    namespace = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    image_media_by_row, cell_image_report = collect_catalog_cell_image_media_paths(
        workbook_path,
        catalog_sheet_names,
    )

    report: dict[str, Any] = {
        "totalAnchoredImages": cell_image_report["totalCellImages"],
        "mappedAnchoredImages": cell_image_report["mappedCellImages"],
        "duplicateAnchorRows": [],
        "missingMediaTargets": [],
        "sheetsWithoutDrawings": [],
        "cellImageReport": cell_image_report,
    }

    with zipfile.ZipFile(workbook_path, "r") as archive:
        sheet_paths = get_workbook_sheet_paths(archive)

        for sheet_name in sorted(catalog_sheet_names):
            sheet_path = sheet_paths.get(sheet_name)

            if not sheet_path:
                report["sheetsWithoutDrawings"].append(sheet_name)
                continue

            sheet_rels = get_relationships(archive, sheet_path)
            drawing_paths = [
                target
                for target in sheet_rels.values()
                if target.startswith("xl/drawings/") and target.endswith(".xml")
            ]

            if not drawing_paths:
                report["sheetsWithoutDrawings"].append(sheet_name)
                continue

            for drawing_path in drawing_paths:
                drawing_root = read_xml(archive, drawing_path)
                drawing_rels = get_relationships(archive, drawing_path)

                anchor_nodes = (
                    drawing_root.findall(".//xdr:twoCellAnchor", namespace)
                    + drawing_root.findall(".//xdr:oneCellAnchor", namespace)
                    + drawing_root.findall(".//xdr:absoluteAnchor", namespace)
                )

                for anchor in anchor_nodes:
                    report["totalAnchoredImages"] += 1

                    from_node = anchor.find("xdr:from", namespace)

                    if from_node is None:
                        continue

                    row_node = from_node.find("xdr:row", namespace)

                    if row_node is None or row_node.text is None:
                        continue

                    row_number = int(row_node.text) + 1

                    blip = anchor.find(".//a:blip", namespace)
                    rel_id = None

                    if blip is not None:
                        rel_id = blip.attrib.get(f"{{{namespace['rel']}}}embed")

                    media_path = drawing_rels.get(rel_id or "")

                    if not media_path:
                        report["missingMediaTargets"].append(
                            {
                                "sheet": sheet_name,
                                "row": row_number,
                                "drawingPath": drawing_path,
                                "relId": rel_id,
                            }
                        )
                        continue

                    key = (sheet_name, row_number)

                    if key in image_media_by_row:
                        report["duplicateAnchorRows"].append(
                            {
                                "sheet": sheet_name,
                                "row": row_number,
                                "firstMediaPath": image_media_by_row[key],
                                "secondMediaPath": media_path,
                            }
                        )
                        continue

                    image_media_by_row[key] = media_path
                    report["mappedAnchoredImages"] += 1

    return image_media_by_row, report

def convert_image_bytes_to_webp(image_bytes: bytes, output_path: Path) -> None:
    with Image.open(BytesIO(image_bytes)) as image:
        # For GIFs, use the first frame. This should be fine for a cover/placeholder.
        try:
            image.seek(0)
        except EOFError:
            pass

        has_transparency = (
            image.mode in {"RGBA", "LA"}
            or "transparency" in image.info
        )

        if has_transparency:
            image = image.convert("RGBA")
        else:
            image = image.convert("RGB")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        image.save(
            output_path,
            "WEBP",
            quality=COVER_WEBP_QUALITY,
            method=6,
        )

def make_cover_public_path(catalog_key: str) -> str:
    return f"{COVER_PUBLIC_PATH}/{catalog_key}.{COVER_IMAGE_EXTENSION}"

def make_cover_output_path(catalog_key: str) -> Path:
    return COVER_OUTPUT_DIR / f"{catalog_key}.{COVER_IMAGE_EXTENSION}"


def hash_image_bytes(image_bytes: bytes) -> str:
    return hashlib.sha256(image_bytes).hexdigest()


def load_cover_cache() -> dict[str, Any]:
    if not COVER_CACHE_PATH.exists():
        return {
            "version": COVER_CACHE_VERSION,
            "covers": {},
        }

    try:
        cache = json.loads(COVER_CACHE_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {
            "version": COVER_CACHE_VERSION,
            "covers": {},
        }

    if cache.get("version") != COVER_CACHE_VERSION:
        return {
            "version": COVER_CACHE_VERSION,
            "covers": {},
        }

    covers = cache.get("covers")

    if not isinstance(covers, dict):
        covers = {}

    return {
        "version": COVER_CACHE_VERSION,
        "covers": covers,
    }


def write_cover_cache(cache: dict[str, Any]) -> None:
    COVER_CACHE_PATH.write_text(
        json.dumps(cache, indent=2, ensure_ascii=False, sort_keys=True),
        encoding="utf-8",
    )

def extract_catalog_cover_images(
    workbook_path: Path,
    catalog_books: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    print("\nExtracting catalog cover images:")

    catalog_sheet_names = {
        str(book["sourceSheet"])
        for book in catalog_books.values()
        if book.get("sourceSheet")
    }

    image_media_by_row, anchor_report = collect_catalog_image_media_paths(
        workbook_path,
        catalog_sheet_names,
    )

    COVER_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    cover_cache = load_cover_cache()
    cached_covers: dict[str, Any] = cover_cache["covers"]
    next_cached_covers: dict[str, Any] = {}

    active_cover_filenames: set[str] = set()

    report: dict[str, Any] = {
        "outputDir": str(COVER_OUTPUT_DIR),
        "publicPath": COVER_PUBLIC_PATH,
        "imageExtension": COVER_IMAGE_EXTENSION,
        "webpQuality": COVER_WEBP_QUALITY,
        "cachePath": str(COVER_CACHE_PATH),
        "extractedCount": 0,
        "convertedCount": 0,
        "cachedCount": 0,
        "deletedStaleCount": 0,
        "missingAnchors": [],
        "missingMediaFiles": [],
        "conversionErrors": [],
        "anchorReport": anchor_report,
    }

    with zipfile.ZipFile(workbook_path, "r") as archive:
        archive_names = set(archive.namelist())
        total_catalog_books = len(catalog_books)

        for index, catalog_book in enumerate(catalog_books.values(), start=1):
            if index == 1 or index % 50 == 0 or index == total_catalog_books:
                percent = index / total_catalog_books * 100
                print(f"  checking covers: {index}/{total_catalog_books} ({percent:.0f}%)")

            source_sheet = catalog_book.get("sourceSheet")
            source_row = catalog_book.get("sourceRow")
            catalog_key = catalog_book["catalogKey"]

            if not source_sheet or source_row is None:
                report["missingAnchors"].append(
                    {
                        "catalogKey": catalog_key,
                        "title": catalog_book["title"],
                        "author": catalog_book["author"],
                        "reason": "missing sourceSheet/sourceRow",
                    }
                )
                catalog_book["coverImage"] = None
                continue

            media_path = image_media_by_row.get((str(source_sheet), int(source_row)))

            if not media_path:
                report["missingAnchors"].append(
                    {
                        "catalogKey": catalog_key,
                        "title": catalog_book["title"],
                        "author": catalog_book["author"],
                        "sourceSheet": source_sheet,
                        "sourceRow": source_row,
                    }
                )
                catalog_book["coverImage"] = None
                continue

            if media_path not in archive_names:
                report["missingMediaFiles"].append(
                    {
                        "catalogKey": catalog_key,
                        "title": catalog_book["title"],
                        "author": catalog_book["author"],
                        "sourceSheet": source_sheet,
                        "sourceRow": source_row,
                        "mediaPath": media_path,
                    }
                )
                catalog_book["coverImage"] = None
                continue

            output_path = make_cover_output_path(catalog_key)
            public_path = make_cover_public_path(catalog_key)
            active_cover_filenames.add(output_path.name)

            try:
                image_bytes = archive.read(media_path)
                image_hash = hash_image_bytes(image_bytes)

                previous_cache_entry = cached_covers.get(catalog_key, {})

                can_reuse_cover = (
                    output_path.exists()
                    and previous_cache_entry.get("imageHash") == image_hash
                    and previous_cache_entry.get("imageExtension") == COVER_IMAGE_EXTENSION
                    and previous_cache_entry.get("webpQuality") == COVER_WEBP_QUALITY
                )

                if can_reuse_cover:
                    report["cachedCount"] += 1
                else:
                    convert_image_bytes_to_webp(image_bytes, output_path)
                    report["convertedCount"] += 1

                catalog_book["coverImage"] = public_path
                report["extractedCount"] += 1

                next_cached_covers[catalog_key] = {
                    "catalogKey": catalog_key,
                    "title": catalog_book["title"],
                    "author": catalog_book["author"],
                    "sourceSheet": source_sheet,
                    "sourceRow": source_row,
                    "mediaPath": media_path,
                    "imageHash": image_hash,
                    "imageExtension": COVER_IMAGE_EXTENSION,
                    "webpQuality": COVER_WEBP_QUALITY,
                    "outputFilename": output_path.name,
                    "publicPath": public_path,
                }

            except (OSError, UnidentifiedImageError, ValueError) as error:
                report["conversionErrors"].append(
                    {
                        "catalogKey": catalog_key,
                        "title": catalog_book["title"],
                        "author": catalog_book["author"],
                        "sourceSheet": source_sheet,
                        "sourceRow": source_row,
                        "mediaPath": media_path,
                        "error": str(error),
                    }
                )
                catalog_book["coverImage"] = None
                continue

    # Delete stale covers after processing, instead of deleting everything upfront.
    for old_cover in COVER_OUTPUT_DIR.glob(f"*.{COVER_IMAGE_EXTENSION}"):
        if old_cover.name not in active_cover_filenames:
            old_cover.unlink()
            report["deletedStaleCount"] += 1

    cover_cache = {
        "version": COVER_CACHE_VERSION,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "covers": next_cached_covers,
    }
    write_cover_cache(cover_cache)

    output_files = list(COVER_OUTPUT_DIR.glob(f"*.{COVER_IMAGE_EXTENSION}"))
    output_bytes = sum(path.stat().st_size for path in output_files)

    report["outputFileCount"] = len(output_files)
    report["outputTotalBytes"] = output_bytes

    print(f"  covers available: {report['extractedCount']}")
    print(f"  converted covers: {report['convertedCount']}")
    print(f"  reused cached covers: {report['cachedCount']}")
    print(f"  deleted stale covers: {report['deletedStaleCount']}")
    print(f"  output files: {report['outputFileCount']}")
    print(f"  output size: {output_bytes / 1024 / 1024:.2f} MB")

    if report["missingAnchors"]:
        print(f"  missing anchors: {len(report['missingAnchors'])}")

    if report["missingMediaFiles"]:
        print(f"  missing media files: {len(report['missingMediaFiles'])}")

    if report["conversionErrors"]:
        print(f"  conversion errors: {len(report['conversionErrors'])}")

    return report

def normalize_book_id(
    value: Any,
    row_number: int,
    title: str,
) -> str:
    book_id = clean(value).lower()

    if not book_id:
        raise ValueError(
            f"List View row {row_number} has a title but no Book ID: "
            f"{title}"
        )

    if not book_id.startswith("book-"):
        raise ValueError(
            f"List View row {row_number} has a malformed Book ID: "
            f"{book_id!r} ({title})"
        )

    guid_text = book_id.removeprefix("book-")

    try:
        parsed_guid = UUID(guid_text)
    except ValueError as error:
        raise ValueError(
            f"List View row {row_number} has a malformed Book ID: "
            f"{book_id!r} ({title})"
        ) from error

    if (
        parsed_guid.version != 4
        or str(parsed_guid) != guid_text
    ):
        raise ValueError(
            f"List View row {row_number} does not contain a canonical "
            f"version-4 Book ID: {book_id!r} ({title})"
        )

    return f"book-{parsed_guid}"

def make_catalog_key(title: str, author_sort: str) -> str:
    slug_source = f"{author_sort}-{title}".lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug_source).strip("-")

    return slug or "untitled"

def load_catalog_books(workbook) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    catalog_books: dict[str, dict[str, Any]] = {}
    processed_sheets: dict[str, int] = {}
    skipped_sheets: list[str] = []
    duplicate_catalog_keys: dict[str, list[dict[str, Any]]] = {}

    for sheet in workbook.worksheets:
        if not is_catalog_sheet(sheet):
            skipped_sheets.append(sheet.title)
            continue

        rows = sheet.iter_rows(values_only=True)

        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration:
            skipped_sheets.append(sheet.title)
            continue

        header_indexes = {
            normalize_header(header): index
            for index, header in enumerate(headers)
        }

        def get(row_values: list[Any], header: str) -> Any:
            index = header_indexes.get(header)
            if index is None or index >= len(row_values):
                return ""
            return row_values[index]

        sheet_count = 0

        for row_number, row in enumerate(rows, start=2):
            row_values = list(row)

            raw_title = clean(get(row_values, "title"))

            if not raw_title:
                continue

            parsed_title = parse_title(raw_title)
            title = clean(parsed_title["title"])
            parsed_series_cell = parse_series_cell(get(row_values, "series"))

            series = parsed_title["series"]
            series_title = parsed_title["seriesTitle"]
            series_format = parsed_title["seriesFormat"]
            series_number = parsed_title["seriesNumber"]

            if parsed_series_cell["series"]:
                series = parsed_series_cell["series"]
                series_title = parsed_series_cell["seriesTitle"]
                series_format = None
                series_number = parsed_series_cell["seriesNumber"]

            first = clean(get(row_values, "first"))
            last = clean(get(row_values, "last"))
            author = make_author(first, last)
            author_sort = make_author_sort(first, last)

            catalog_key = make_catalog_key(title, author_sort)

            catalog_book = {
                "catalogKey": catalog_key,
                "title": title,
                "rawTitle": raw_title,
                "coverImage": None,
                "series": series,
                "seriesTitle": series_title,
                "seriesFormat": series_format,
                "seriesNumber": series_number,
                "author": author,
                "authorSort": author_sort,
                "firstName": first,
                "lastName": last,
                "format": clean(get(row_values, "format")),
                "jc": checkbox_to_bool(get(row_values, "jc")),
                "cj": checkbox_to_bool(get(row_values, "cj")),
                "lgbtq": checkbox_to_bool(get(row_values, "lgbtq+")),
                "sourceSheet": sheet.title,
                "sourceRow": row_number,
            }

            if catalog_key in catalog_books:
                duplicate_catalog_keys.setdefault(catalog_key, [catalog_books[catalog_key]])
                duplicate_catalog_keys[catalog_key].append(catalog_book)

            catalog_books[catalog_key] = catalog_book
            sheet_count += 1

        processed_sheets[sheet.title] = sheet_count

    report = {
        "catalogBookCount": len(catalog_books),
        "catalogRowsWithTitles": sum(processed_sheets.values()),
        "processedCatalogSheets": processed_sheets,
        "skippedCatalogSheets": skipped_sheets,
        "duplicateCatalogKeys": duplicate_catalog_keys,
    }

    return catalog_books, report

def make_wanted_id(index: int, list_type: str, title: str, author_sort: str, series: str | None) -> str:
    slug_source = f"{list_type}-{author_sort}-{series or ''}-{title}".lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug_source).strip("-")
    slug = slug[:70].strip("-")

    return f"wanted-{list_type}-{index:04d}-{slug or 'untitled'}"

def load_wanted_sheet(
    workbook,
    sheet_name: str,
    list_type: str,
    series_headers: tuple[str, ...] = (),
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return [], {
            "sheetName": sheet_name,
            "found": False,
            "headers": [],
            "rowCount": 0,
            "skippedBlankRows": 0,
        }

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)

    try:
        first_row = next(rows)
    except StopIteration:
        return [], {
            "sheetName": sheet_name,
            "found": True,
            "headers": [],
            "headerMode": "empty",
            "rowCount": 0,
            "skippedBlankRows": 0,
        }

    headers = [clean(value) for value in first_row]
    header_indexes = {
        normalize_header(header): index
        for index, header in enumerate(headers)
    }
    header_mode = "headers"
    row_iterator = list(enumerate(rows, start=2))

    # The bookstore sheets started life as quick headerless notes. Prefer headers
    # when present, but keep the fallback so an older workbook copy does not
    # make the whole build explode.
    if get_header_index(header_indexes, "title", "booktitle") is None:
        if series_headers:
            headers = ["Title", "Series", "First", "Last"]
        else:
            headers = ["Title", "First", "Last"]

        header_indexes = {
            normalize_header(header): index
            for index, header in enumerate(headers)
        }
        header_mode = "fallback-default"
        row_iterator = [(1, first_row)] + row_iterator

    wanted_books: list[dict[str, Any]] = []
    skipped_blank_rows = 0

    for row_number, row in row_iterator:
        row_values = list(row)

        raw_title = clean(
            get_value_by_header_alias(row_values, header_indexes, "title", "booktitle")
        )

        if not raw_title:
            skipped_blank_rows += 1
            continue

        first = clean(
            get_value_by_header_alias(
                row_values,
                header_indexes,
                "first",
                "firstname",
                "authorfirst",
                "authorfirstname",
                "author first",
                "author first name",
            )
        )
        last = clean(
            get_value_by_header_alias(
                row_values,
                header_indexes,
                "last",
                "lastname",
                "authorlast",
                "authorlastname",
                "author last",
                "author last name",
            )
        )

        author = make_author(first, last)
        author_sort = make_author_sort(first, last)
        parsed_title = parse_title(raw_title)
        title = clean(parsed_title["title"])

        raw_series = clean(
            get_value_by_header_alias(row_values, header_indexes, *series_headers)
        ) if series_headers else ""
        parsed_series_cell = parse_series_cell(raw_series)
        series = parsed_series_cell["series"]
        series_title = parsed_series_cell["seriesTitle"]
        series_number = parsed_series_cell["seriesNumber"]

        wanted_books.append(
            {
                "wantedId": make_wanted_id(
                    len(wanted_books) + 1,
                    list_type,
                    title,
                    author_sort,
                    series if isinstance(series, str) else None,
                ),
                "listType": list_type,
                "title": title,
                "rawTitle": raw_title,
                "series": series,
                "seriesTitle": series_title,
                "seriesNumber": series_number,
                "author": author,
                "authorSort": author_sort,
                "firstName": first,
                "lastName": last,
                "sourceSheet": sheet_name,
                "sourceRow": row_number,
            }
        )

    report = {
        "sheetName": sheet_name,
        "found": True,
        "headers": headers,
        "headerMode": header_mode,
        "rowCount": len(wanted_books),
        "skippedBlankRows": skipped_blank_rows,
    }

    return wanted_books, report

def load_wanted_lists(workbook) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    to_buy, to_buy_report = load_wanted_sheet(
        workbook,
        sheet_name="To Buy",
        list_type="to-buy",
    )
    series_to_complete, series_to_complete_report = load_wanted_sheet(
        workbook,
        sheet_name="Series to Complete",
        list_type="series-to-complete",
        series_headers=(
            "series",
            "series name",
            "series name and number",
            "series name and number in the series",
        ),
    )

    wanted_lists = {
        "toBuy": to_buy,
        "seriesToComplete": series_to_complete,
    }

    report = {
        "toBuy": to_buy_report,
        "seriesToComplete": series_to_complete_report,
        "totalWantedCount": len(to_buy) + len(series_to_complete),
    }

    return wanted_lists, report

def normalize_match_text(value: Any) -> str:
    text = clean(value).lower()
    text = (
        text.replace("’", "'")
        .replace("‘", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("–", "-")
        .replace("—", "-")
    )
    text = re.sub(r"&", " and ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()

def is_short_title_match(short_title: str, full_title: str) -> bool:
    short_text = normalize_match_text(short_title)
    full_text = normalize_match_text(full_title)

    if not short_text or not full_text:
        return False

    if short_text == full_text:
        return True

    # Allows List View to use a shorter display title while catalog keeps
    # full subtitle / packaging text, e.g.
    # "The Hundred Years' War on Palestine"
    # -> "The Hundred Years' War on Palestine: A History..."
    return full_text.startswith(f"{short_text} ")

def find_catalog_match(
    catalog_books: dict[str, dict[str, Any]],
    title: str,
    author_sort: str,
) -> tuple[dict[str, Any] | None, str]:
    exact_catalog_key = make_catalog_key(title, author_sort)
    exact_match = catalog_books.get(exact_catalog_key)

    if exact_match:
        return exact_match, "exact"

    normalized_author_sort = normalize_match_text(author_sort)

    title_prefix_matches = [
        catalog_book
        for catalog_book in catalog_books.values()
        if normalize_match_text(catalog_book["authorSort"]) == normalized_author_sort
        and is_short_title_match(title, catalog_book["title"])
    ]

    if len(title_prefix_matches) == 1:
        return title_prefix_matches[0], "title-prefix"

    return None, "missing"

def main() -> None:
    for workbook_path in (
        CATALOG_WORKBOOK_PATH,
        LIST_VIEW_WORKBOOK_PATH,
    ):
        if not workbook_path.exists():
            raise FileNotFoundError(
                f"Could not find workbook: "
                f"{workbook_path}"
            )

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    print(
        f"Loading Catalog workbook: "
        f"{CATALOG_WORKBOOK_PATH}"
    )

    catalog_workbook = load_workbook(
        CATALOG_WORKBOOK_PATH,
        data_only=True,
    )

    print(
        f"Loading List View workbook: "
        f"{LIST_VIEW_WORKBOOK_PATH}"
    )

    list_view_workbook = load_workbook(
        LIST_VIEW_WORKBOOK_PATH,
        data_only=True,
    )

    bookcase_rooms = load_bookcase_rooms(
        catalog_workbook
    )

    catalog_books, catalog_report = (
        load_catalog_books(
            catalog_workbook
        )
    )

    if not catalog_books:
        raise ValueError(
            "Catalog safety stop: no Catalog books were detected. "
            "The build was aborted before generated covers or JSON files "
            "could be deleted or replaced."
        )

    wanted_lists, wanted_report = (
        load_wanted_lists(
            catalog_workbook
        )
    )

    cover_extraction_report = (
        extract_catalog_cover_images(
            CATALOG_WORKBOOK_PATH,
            catalog_books,
        )
    )

    if DEBUG_IMAGE_INSPECTION:
        image_report = inspect_catalog_images(catalog_workbook)
        image_package_report = inspect_workbook_image_package(CATALOG_WORKBOOK_PATH)

        catalog_sheet_names = {
            sheet.title
            for sheet in catalog_workbook.worksheets
            if is_catalog_sheet(sheet)
        }

        drawing_anchor_report = inspect_drawing_image_anchors(
            CATALOG_WORKBOOK_PATH,
            catalog_sheet_names,
            catalog_books,
        )
    else:
        image_report = None
        image_package_report = None
        drawing_anchor_report = None

    print(f"Found {len(catalog_books)} catalog books")

    print("\nProcessed catalog sheets:")
    for sheet_name, count in catalog_report["processedCatalogSheets"].items():
        print(f"  - {sheet_name}: {count}")

    print("\nSkipped non-catalog sheets:")
    for sheet_name in catalog_report["skippedCatalogSheets"]:
        print(f"  - {sheet_name}")

    if catalog_report["duplicateCatalogKeys"]:
        print("\nDuplicate catalog keys:")
        for catalog_key, duplicates in catalog_report["duplicateCatalogKeys"].items():
            print(f"  - {catalog_key}: {len(duplicates)} rows")

    sheet = find_sheet_with_headers(
        list_view_workbook,
        required_headers={
            "cj",
            "jc",
            "lgbtq+",
            "isbn",
            "year",
            "pages",
            "title",
            "series",
            "first",
            "last",
            "genre",
            "subgenre",
            "publisher",
            "origin",
            "bookcase",
            "shelf",
            "position",
            "systemcolumns-automationonly",
            "bookid",
        },
        label="List View",
    )
    print(f"Using List View sheet: {sheet.title}")

    rows = sheet.iter_rows(values_only=True)

    try:
        headers = [clean(value) for value in next(rows)]
    except StopIteration:
        raise ValueError("The List View sheet is empty.")

    header_indexes = {normalize_header(header): index for index, header in enumerate(headers)}

    required_headers = [
        "cj",
        "jc",
        "lgbtq+",
        "isbn",
        "year",
        "pages",
        "title",
        "series",
        "first",
        "last",
        "genre",
        "subgenre",
        "publisher",
        "origin",
        "bookcase",
        "shelf",
        "position",
        "systemcolumns-automationonly",
        "bookid",
    ]

    missing_headers = [
        header
        for header in required_headers
        if header not in header_indexes
    ]

    if missing_headers:
        raise ValueError(
            "Missing expected headers: "
            + ", ".join(missing_headers)
            + f"\nFound headers: {headers}"
        )

    boundary_header_index = header_indexes[
        "systemcolumns-automationonly"
    ]

    book_id_header_index = header_indexes[
        "bookid"
    ]

    if (
        headers[boundary_header_index]
        != "SYSTEM COLUMNS - AUTOMATION ONLY"
    ):
        raise ValueError(
            "The List View automation boundary header does not "
            "exactly match: SYSTEM COLUMNS - AUTOMATION ONLY"
        )

    if headers[book_id_header_index] != "Book ID":
        raise ValueError(
            "The List View Book ID header does not exactly "
            "match: Book ID"
        )

    if book_id_header_index != boundary_header_index + 1:
        raise ValueError(
            "Book ID must be immediately after "
            "SYSTEM COLUMNS - AUTOMATION ONLY."
        )

    books = []
    skipped_blank_rows = 0
    unmapped_bookcases = set()
    used_bookcases = set()
    catalog_match_fallback_rows = []
    catalog_unmatched_rows = []
    book_id_rows: dict[str, int] = {}

    for row_number, row in enumerate(rows, start=2):
        row_values = list(row)

        def get(header: str) -> str:
            index = header_indexes.get(header)
            if index is None or index >= len(row_values):
                return ""
            return clean(row_values[index])

        raw_title = get("title")

        if not raw_title:
            skipped_blank_rows += 1
            continue

        book_id = normalize_book_id(
            get("bookid"),
            row_number,
            raw_title,
        )

        existing_book_id_row = book_id_rows.get(
            book_id
        )

        if existing_book_id_row is not None:
            raise ValueError(
                f"Duplicate Book ID {book_id!r} found in "
                f"List View rows {existing_book_id_row} "
                f"and {row_number}."
            )

        book_id_rows[book_id] = row_number

        parsed_title = parse_title(raw_title)
        title = clean(parsed_title["title"])
        parsed_series_cell = parse_series_cell(get("series"))

        series = parsed_title["series"]
        series_title = parsed_title["seriesTitle"]
        series_format = parsed_title["seriesFormat"]
        series_number = parsed_title["seriesNumber"]

        if parsed_series_cell["series"]:
            series = parsed_series_cell["series"]
            series_title = parsed_series_cell["seriesTitle"]
            series_format = None
            series_number = parsed_series_cell["seriesNumber"]

        first = get("first")
        last = get("last")
        author = make_author(first, last)
        author_sort = make_author_sort(first, last)

        bookcase = get("bookcase")
        room = get_room_for_bookcase(bookcase, bookcase_rooms)

        catalog_key = make_catalog_key(title, author_sort)
        catalog_match, catalog_match_type = find_catalog_match(
            catalog_books,
            title,
            author_sort,
        )

        if catalog_match:
            catalog_key = catalog_match["catalogKey"]

            if catalog_match_type != "exact":
                catalog_match_fallback_rows.append(
                    {
                        "row": row_number,
                        "matchType": catalog_match_type,
                        "listTitle": title,
                        "catalogTitle": catalog_match["title"],
                        "author": author,
                    }
                )
        else:
            catalog_unmatched_rows.append(
                {
                    "row": row_number,
                    "title": title,
                    "author": author,
                    "catalogKey": catalog_key,
                }
            )

        if bookcase:
            used_bookcases.add(bookcase)

        if bookcase and not room:
            unmapped_bookcases.add(bookcase)

        raw_shelf = get("shelf")
        shelf, row_name = parse_shelf(raw_shelf)

        cj = checkbox_to_bool(
            get("cj")
        )

        jc = checkbox_to_bool(
            get("jc")
        )

        lgbtq = checkbox_to_bool(
            get("lgbtq+")
        )

        if catalog_match:
            catalog_match["cj"] = cj
            catalog_match["jc"] = jc
            catalog_match["lgbtq"] = (
                lgbtq
            )

        book = {
            "bookId": book_id,
            "title": title,
            "rawTitle": raw_title,
            "catalogTitle": (
                catalog_match["title"]
                if catalog_match
                else title
            ),
            "catalogRawTitle": (
                catalog_match["rawTitle"]
                if catalog_match
                else raw_title
            ),
            "catalogMatchType": catalog_match_type,
            "shelfPosition": parse_optional_int(
                get("position")
            ),
            "series": series,
            "seriesTitle": series_title,
            "seriesFormat": series_format,
            "seriesNumber": series_number,
            "author": author,
            "authorSort": author_sort,
            "firstName": first,
            "lastName": last,
            "genre": get("genre"),
            "subgenre": get("subgenre"),
            "publisher": get("publisher"),
            "origin": get("origin"),
            "isbn": get("isbn"),
            "publicationYear": parse_optional_int(
                get("year")
            ),
            "totalPages": parse_optional_int(
                get("pages")
            ),
            "catalogKey": catalog_key,
            "format": (
                catalog_match["format"]
                if catalog_match
                else ""
            ),
            "jc": jc,
            "cj": cj,
            "lgbtq": lgbtq,
            "coverImage": (
                catalog_match.get("coverImage")
                if catalog_match
                else None
            ),
            "room": room,
            "bookcase": bookcase,
            "shelf": shelf,
            "row": row_name,
            "rawShelf": raw_shelf,
            "notes": "",
        }

        books.append(book)

    unusedBookcases = sorted(
        bookcase for bookcase in bookcase_rooms
        if bookcase not in used_bookcases
    )

    meta = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": str(
            LIST_VIEW_WORKBOOK_PATH
        ),

        "catalogWorkbook": str(
            CATALOG_WORKBOOK_PATH
        ),
        "sourceSheet": sheet.title,
        "bookCount": len(books),
        "bookIdCount": len(book_id_rows),
        "bookIdSource": "List View.Book ID",
        "skippedBlankRows": skipped_blank_rows,
        "headers": headers,
        "bookcaseRooms": bookcase_rooms,
        "usedBookcases": sorted(used_bookcases),
        "unusedBookcases": unusedBookcases,
        "unmappedBookcases": sorted(unmapped_bookcases),
        "catalogMatchFallbackRows": catalog_match_fallback_rows,
        "catalogUnmatchedRows": catalog_unmatched_rows,
        "catalog": catalog_report,
        "wanted": wanted_report,
        "coverExtraction": cover_extraction_report,
        "imageInspectionEnabled": DEBUG_IMAGE_INSPECTION,
        "images": image_report,
        "imagePackage": image_package_report,
        "drawingAnchors": drawing_anchor_report,
    }

    BOOKS_OUTPUT_PATH.write_text(
        json.dumps(books, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    WANTED_OUTPUT_PATH.write_text(
        json.dumps(wanted_lists, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    CATALOG_OUTPUT_PATH.write_text(
        json.dumps(
            sorted(
                catalog_books.values(),
                key=lambda book: (
                    book["authorSort"].lower(),
                    book["title"].lower(),
                    book["sourceSheet"],
                    book["sourceRow"],
                ),
            ),
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    META_OUTPUT_PATH.write_text(
        json.dumps(meta, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"Wrote {len(books)} books to {BOOKS_OUTPUT_PATH}")
    print(
        f"Wrote {wanted_report['totalWantedCount']} wanted books "
        f"to {WANTED_OUTPUT_PATH}"
    )
    print(f"Wrote {len(catalog_books)} catalog books to {CATALOG_OUTPUT_PATH}")
    print(f"Wrote metadata to {META_OUTPUT_PATH}")

    if catalog_match_fallback_rows:
        print("\nCatalog title-prefix matches:")
        for match in catalog_match_fallback_rows:
            print(
                f"  - row {match['row']}: "
                f"{match['listTitle']} -> {match['catalogTitle']}"
            )

    if catalog_unmatched_rows:
        print("\nRows without catalog matches:")
        for unmatched in catalog_unmatched_rows:
            print(
                f"  - row {unmatched['row']}: "
                f"{unmatched['title']} by {unmatched['author']}"
            )

    if skipped_blank_rows:
        print(f"Skipped {skipped_blank_rows} blank rows")

    if unmapped_bookcases:
        print("\nUnmapped bookcases:")
        for bookcase in sorted(unmapped_bookcases):
            print(f"  - {bookcase}")

    catalog_workbook.close()
    list_view_workbook.close()


if __name__ == "__main__":
    main()