#!/usr/bin/env python3
"""Export archived/offloaded library records to a personal review workbook."""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_OUTPUT_PATH = Path("C:/library_app/MyLibrary Archive/ARCHIVE.xlsx")
DEFAULT_ARCHIVE_PATH = (
    Path(__file__).resolve().parents[1]
    / "public"
    / "data"
    / "library-archive.json"
)
DEFAULT_SOURCE_COMMIT = "2c3d3c7"

COLUMNS = [
    ("archiveStatus", "Archive Status"),
    ("offloadDate", "Offload Date"),
    ("offloadDestination", "Offload Destination"),
    ("title", "Title"),
    ("author", "Author"),
    ("firstName", "First"),
    ("lastName", "Last"),
    ("seriesTitle", "Series"),
    ("seriesNumber", "Series #"),
    ("genre", "Genre"),
    ("subgenre", "Subgenre"),
    ("publisher", "Publisher"),
    ("format", "Format"),
    ("isbn", "ISBN"),
    ("publicationYear", "Publication Year"),
    ("totalPages", "Total Pages"),
    ("origin", "Origin"),
    ("cj", "CJ Read"),
    ("jc", "JC Read"),
    ("bipoc", "BIPOC"),
    ("lgbtq", "LGBTQ+"),
    ("lastKnownRoom", "Last Known Room"),
    ("lastKnownBookcase", "Last Known Bookcase"),
    ("lastKnownShelf", "Last Known Shelf"),
    ("lastKnownRow", "Last Known Row"),
    ("rawShelf", "Raw Shelf"),
    ("coverFile", "Local Cover File"),
    ("coverImage", "App Cover Path"),
    ("bookId", "Book ID"),
    ("catalogKey", "Catalog Key"),
    ("notes", "Original Notes"),
    ("archiveNotes", "Archive Notes"),
]


def load_archive(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    if not isinstance(data, list):
        raise ValueError(f"Expected a JSON array in {path}")

    records: list[dict[str, Any]] = []

    for index, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"Archive item {index} is not an object")

        records.append(item)

    return records


def cover_relative_path(record: dict[str, Any]) -> Path | None:
    cover_image = str(record.get("coverImage") or "").strip()

    if not cover_image.startswith("/data/covers/"):
        return None

    return Path(*cover_image.lstrip("/").split("/"))


def copy_cover(
    record: dict[str, Any],
    repo_root: Path,
    covers_dir: Path,
    source_commit: str,
) -> str:
    relative_cover = cover_relative_path(record)

    if relative_cover is None:
        return ""

    source_path = repo_root / "public" / relative_cover
    destination_path = covers_dir / relative_cover.name
    destination_path.parent.mkdir(parents=True, exist_ok=True)

    if source_path.exists():
        shutil.copy2(source_path, destination_path)
        return str(destination_path)

    git_path = f"public/{relative_cover.as_posix()}"

    try:
        cover_bytes = subprocess.check_output(
            [
                "git",
                "show",
                f"{source_commit}:{git_path}",
            ],
            cwd=repo_root,
        )
    except (
        subprocess.CalledProcessError,
        FileNotFoundError,
    ):
        return ""

    destination_path.write_bytes(cover_bytes)
    return str(destination_path)


def bool_display(value: Any) -> str:
    if value is True:
        return "TRUE"

    if value is False:
        return "FALSE"

    return ""


def cell_value(record: dict[str, Any], key: str) -> Any:
    if key in {"cj", "jc", "bipoc", "lgbtq"}:
        return bool_display(record.get(key))

    return record.get(key)


def write_workbook(
    records: list[dict[str, Any]],
    output_path: Path,
    repo_root: Path,
    source_commit: str,
) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    covers_dir = output_path.parent / "covers"

    records = sorted(
        records,
        key=lambda record: (
            str(record.get("authorSort") or ""),
            str(record.get("title") or ""),
        ),
    )

    copied_covers = 0

    for record in records:
        cover_file = copy_cover(
            record,
            repo_root,
            covers_dir,
            source_commit,
        )

        record["coverFile"] = cover_file

        if cover_file:
            copied_covers += 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Archive"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="E8EEF7",
    )

    sheet.append([label for _, label in COLUMNS])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top")

    for record in records:
        sheet.append(
            [
                cell_value(record, key)
                for key, _ in COLUMNS
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 18,
        "B": 14,
        "C": 22,
        "D": 42,
        "E": 28,
        "F": 18,
        "G": 18,
        "H": 28,
        "I": 10,
        "J": 20,
        "K": 22,
        "L": 26,
        "M": 16,
        "N": 18,
        "O": 16,
        "P": 12,
        "Q": 16,
        "R": 10,
        "S": 10,
        "T": 10,
        "U": 10,
        "V": 18,
        "W": 22,
        "X": 18,
        "Y": 14,
        "Z": 18,
        "AA": 44,
        "AB": 40,
        "AC": 40,
        "AD": 42,
        "AE": 36,
        "AF": 36,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    cover_column = next(
        index + 1
        for index, (key, _) in enumerate(COLUMNS)
        if key == "coverFile"
    )

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=cover_column)

        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"

    workbook.save(output_path)
    return copied_covers


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export MyLibrary archive records to ARCHIVE.xlsx"
    )

    parser.add_argument(
        "--archive",
        type=Path,
        default=DEFAULT_ARCHIVE_PATH,
        help="Path to library-archive.json",
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="Output ARCHIVE.xlsx path",
    )

    parser.add_argument(
        "--source-commit",
        default=DEFAULT_SOURCE_COMMIT,
        help="Git commit to recover missing cover files from",
    )

    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]

    try:
        records = load_archive(args.archive)
        copied_covers = write_workbook(
            records,
            args.output,
            repo_root,
            args.source_commit,
        )
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1

    print(f"Archived records: {len(records)}")
    print(f"Cover files available: {copied_covers}")
    print(f"Wrote: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
