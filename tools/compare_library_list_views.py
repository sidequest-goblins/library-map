from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "\nERROR: openpyxl is not installed.\n"
        "Run this inside the library-map virtual environment:\n\n"
        "    python -m pip install openpyxl\n"
    )
    raise SystemExit(1)


DEFAULT_FOLDER = Path(
    r"C:\Users\cjade\OneDrive\Shared Workbooks\MyLibrary"
)
DEFAULT_REGULAR = DEFAULT_FOLDER / "LIBRARY LIST VIEW.xlsx"
DEFAULT_COPY = DEFAULT_FOLDER / "LIBRARY LIST VIEW - Copy.xlsx"
DEFAULT_REPORT = DEFAULT_FOLDER / "LIBRARY LIST VIEW COMPARISON.xlsx"

IGNORE_COLUMN_NUMBER = 1  # Entire first column, CJ, is ignored.
HEADER_SEARCH_LIMIT = 30


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().casefold().split())


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def display_value(value: Any) -> Any:
    """Convert values into report-safe, readable Excel values."""
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, time):
        return value.strftime("%H:%M:%S")

    if isinstance(value, str):
        # Keep literal text from being interpreted as a formula in the report.
        if value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    return value


def comparable_value(value: Any) -> tuple[str, Any]:
    """
    Normalize values enough to avoid fake differences such as 5 versus 5.0,
    while preserving real differences such as False versus 0.
    """
    if is_blank(value):
        return ("blank", None)

    if isinstance(value, bool):
        return ("bool", value)

    if isinstance(value, (int, float, Decimal)):
        try:
            return ("number", Decimal(str(value)).normalize())
        except InvalidOperation:
            return ("other", str(value))

    if isinstance(value, datetime):
        return ("datetime", value.replace(microsecond=0).isoformat())

    if isinstance(value, date):
        return ("date", value.isoformat())

    if isinstance(value, time):
        return ("time", value.replace(microsecond=0).isoformat())

    if isinstance(value, str):
        return ("string", value)

    return ("other", str(value))


def values_equal(left: Any, right: Any) -> bool:
    return comparable_value(left) == comparable_value(right)


def classify_change(left: Any, right: Any) -> str:
    if is_blank(left) and not is_blank(right):
        return "Filled in on Copy"

    if not is_blank(left) and is_blank(right):
        return "Blank on Copy"

    if (
        isinstance(left, str)
        and isinstance(right, str)
        and left != right
        and left.strip() == right.strip()
    ):
        return "Whitespace only"

    return "Changed"


def choose_sheet(workbook):
    if "List View" in workbook.sheetnames:
        return workbook["List View"]

    return workbook.active


def find_header_row_and_columns(worksheet) -> tuple[int, dict[str, int], dict[str, str]]:
    """
    Locate the header row by looking for Book ID.

    Returns:
        header row number
        normalized header -> column number
        normalized header -> original display header
    """
    max_search_row = min(worksheet.max_row, HEADER_SEARCH_LIMIT)

    for row_number in range(1, max_search_row + 1):
        row_headers: dict[str, int] = {}
        display_headers: dict[str, str] = {}

        for column_number in range(1, worksheet.max_column + 1):
            if column_number == IGNORE_COLUMN_NUMBER:
                continue

            raw_header = worksheet.cell(row_number, column_number).value
            normalized = normalize_header(raw_header)

            if not normalized:
                continue

            if normalized not in row_headers:
                row_headers[normalized] = column_number
                display_headers[normalized] = str(raw_header).strip()

        if "book id" in row_headers:
            return row_number, row_headers, display_headers

    raise ValueError(
        f'Could not find a header row containing "Book ID" '
        f'in sheet "{worksheet.title}".'
    )


def row_has_data(
    worksheet,
    row_number: int,
    header_columns: dict[str, int],
) -> bool:
    return any(
        not is_blank(worksheet.cell(row_number, column_number).value)
        for column_number in header_columns.values()
    )


def collect_rows(
    worksheet,
    header_row: int,
    header_columns: dict[str, int],
) -> tuple[dict[str, int], dict[str, list[int]], list[int]]:
    """
    Returns:
        unique Book ID -> row number
        duplicate Book ID -> all affected row numbers
        nonblank rows without Book ID
    """
    book_id_column = header_columns["book id"]

    first_seen: dict[str, int] = {}
    duplicate_rows: dict[str, list[int]] = {}
    unkeyed_rows: list[int] = []

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        if not row_has_data(worksheet, row_number, header_columns):
            continue

        raw_book_id = worksheet.cell(row_number, book_id_column).value

        if is_blank(raw_book_id):
            unkeyed_rows.append(row_number)
            continue

        book_id = str(raw_book_id).strip()

        if book_id in first_seen:
            duplicate_rows.setdefault(
                book_id,
                [first_seen[book_id]],
            ).append(row_number)
        else:
            first_seen[book_id] = row_number

    # Duplicate IDs cannot safely be used for matching.
    for duplicate_id in duplicate_rows:
        first_seen.pop(duplicate_id, None)

    return first_seen, duplicate_rows, unkeyed_rows


def get_value(
    worksheet,
    row_number: int,
    header_columns: dict[str, int],
    normalized_header: str,
) -> Any:
    column_number = header_columns.get(normalized_header)

    if column_number is None:
        return None

    return worksheet.cell(row_number, column_number).value


def book_context(
    regular_sheet,
    copy_sheet,
    regular_row: int | None,
    copy_row: int | None,
    regular_headers: dict[str, int],
    copy_headers: dict[str, int],
) -> tuple[Any, str]:
    title = None
    first_name = None
    last_name = None

    if regular_row is not None:
        title = get_value(
            regular_sheet, regular_row, regular_headers, "title"
        )
        first_name = get_value(
            regular_sheet, regular_row, regular_headers, "first"
        )
        last_name = get_value(
            regular_sheet, regular_row, regular_headers, "last"
        )

    if copy_row is not None:
        if is_blank(title):
            title = get_value(copy_sheet, copy_row, copy_headers, "title")
        if is_blank(first_name):
            first_name = get_value(
                copy_sheet, copy_row, copy_headers, "first"
            )
        if is_blank(last_name):
            last_name = get_value(
                copy_sheet, copy_row, copy_headers, "last"
            )

    author = " ".join(
        str(part).strip()
        for part in (first_name, last_name)
        if not is_blank(part)
    )

    return title, author


def sheet_settings(worksheet) -> dict[str, str]:
    settings: dict[str, str] = {
        "Sheet name": worksheet.title,
        "Maximum populated row": str(worksheet.max_row),
        "Maximum populated column": str(worksheet.max_column),
        "Freeze panes": str(worksheet.freeze_panes or ""),
        "AutoFilter range": str(worksheet.auto_filter.ref or ""),
        "Merged ranges": ", ".join(
            str(item) for item in worksheet.merged_cells.ranges
        ),
        "Show gridlines": str(worksheet.sheet_view.showGridLines),
    }

    for column_number in range(2, worksheet.max_column + 1):
        letter = get_column_letter(column_number)
        dimension = worksheet.column_dimensions[letter]

        if dimension.hidden:
            settings[f"Column {letter} hidden"] = "True"

        if dimension.width is not None:
            settings[f"Column {letter} width"] = str(dimension.width)

    hidden_rows = [
        str(row_number)
        for row_number, dimension in worksheet.row_dimensions.items()
        if dimension.hidden
    ]
    settings["Hidden rows"] = ", ".join(hidden_rows)

    return settings


def make_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)

    fill = PatternFill("solid", fgColor="4F81BD")
    font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def autosize(sheet, maximum_width: int = 55) -> None:
    for column_cells in sheet.columns:
        longest = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            longest = max(longest, len(text))

        sheet.column_dimensions[column_letter].width = min(
            max(longest + 2, 10),
            maximum_width,
        )

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def write_full_rows_sheet(
    output_workbook: Workbook,
    sheet_name: str,
    source_sheet,
    book_ids: list[str],
    row_lookup: dict[str, int],
    header_columns: dict[str, int],
    display_headers: dict[str, str],
) -> None:
    sheet = output_workbook.create_sheet(sheet_name)

    ordered_headers = sorted(
        header_columns,
        key=lambda item: header_columns[item],
    )

    headers = ["Source Excel row"] + [
        display_headers[item] for item in ordered_headers
    ]
    make_header(sheet, headers)

    for book_id in book_ids:
        row_number = row_lookup[book_id]
        values = [row_number]

        for header in ordered_headers:
            value = source_sheet.cell(
                row_number,
                header_columns[header],
            ).value
            values.append(display_value(value))

        sheet.append(values)

    autosize(sheet)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Compare two Library List View workbooks while ignoring "
            "their entire first column."
        )
    )
    parser.add_argument(
        "regular",
        nargs="?",
        type=Path,
        default=DEFAULT_REGULAR,
        help=f"Regular workbook. Default: {DEFAULT_REGULAR}",
    )
    parser.add_argument(
        "copy",
        nargs="?",
        type=Path,
        default=DEFAULT_COPY,
        help=f"Copy workbook. Default: {DEFAULT_COPY}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_REPORT,
        help=f"Comparison report. Default: {DEFAULT_REPORT}",
    )
    args = parser.parse_args()

    regular_path = args.regular.expanduser().resolve()
    copy_path = args.copy.expanduser().resolve()
    output_path = args.output.expanduser().resolve()

    missing_files = [
        path for path in (regular_path, copy_path) if not path.exists()
    ]

    if missing_files:
        print("\nERROR: Could not find:")
        for path in missing_files:
            print(f"  - {path}")
        return 1

    if regular_path == copy_path:
        print("\nERROR: The two input paths point to the same file.")
        return 1

    print("\nLoading workbooks in read-only comparison mode...")
    print(f"Regular: {regular_path}")
    print(f"Copy:    {copy_path}")

    try:
        # data_only=True compares Excel's last saved/displayed formula results,
        # avoiding row-reference formula noise after sorting.
        regular_workbook = load_workbook(
            regular_path,
            data_only=True,
            read_only=False,
        )
        copy_workbook = load_workbook(
            copy_path,
            data_only=True,
            read_only=False,
        )
    except PermissionError:
        print(
            "\nERROR: Windows would not allow one of the workbooks to be read.\n"
            "Save and close both workbooks in Excel, then run the script again."
        )
        return 1
    except Exception as error:
        print(f"\nERROR while opening the workbooks: {error}")
        return 1

    regular_sheet = choose_sheet(regular_workbook)
    copy_sheet = choose_sheet(copy_workbook)

    try:
        (
            regular_header_row,
            regular_headers,
            regular_display_headers,
        ) = find_header_row_and_columns(regular_sheet)

        (
            copy_header_row,
            copy_headers,
            copy_display_headers,
        ) = find_header_row_and_columns(copy_sheet)
    except ValueError as error:
        print(f"\nERROR: {error}")
        return 1

    (
        regular_rows,
        regular_duplicates,
        regular_unkeyed,
    ) = collect_rows(
        regular_sheet,
        regular_header_row,
        regular_headers,
    )

    (
        copy_rows,
        copy_duplicates,
        copy_unkeyed,
    ) = collect_rows(
        copy_sheet,
        copy_header_row,
        copy_headers,
    )

    regular_header_names = set(regular_headers)
    copy_header_names = set(copy_headers)

    common_headers = regular_header_names & copy_header_names
    regular_only_headers = regular_header_names - copy_header_names
    copy_only_headers = copy_header_names - regular_header_names

    common_book_ids = set(regular_rows) & set(copy_rows)
    regular_only_book_ids = sorted(set(regular_rows) - set(copy_rows))
    copy_only_book_ids = sorted(set(copy_rows) - set(regular_rows))

    ordered_common_headers = sorted(
        common_headers,
        key=lambda item: regular_headers[item],
    )

    # Book ID is used as the match key, so there is no need to report it as
    # a changed data field.
    comparable_headers = [
        item for item in ordered_common_headers if item != "book id"
    ]

    differences: list[list[Any]] = []

    for book_id in sorted(common_book_ids):
        regular_row = regular_rows[book_id]
        copy_row = copy_rows[book_id]

        title, author = book_context(
            regular_sheet,
            copy_sheet,
            regular_row,
            copy_row,
            regular_headers,
            copy_headers,
        )

        for normalized_header in comparable_headers:
            regular_value = regular_sheet.cell(
                regular_row,
                regular_headers[normalized_header],
            ).value
            copy_value = copy_sheet.cell(
                copy_row,
                copy_headers[normalized_header],
            ).value

            if values_equal(regular_value, copy_value):
                continue

            differences.append(
                [
                    book_id,
                    display_value(title),
                    author,
                    regular_display_headers.get(
                        normalized_header,
                        normalized_header,
                    ),
                    regular_row,
                    copy_row,
                    display_value(regular_value),
                    display_value(copy_value),
                    classify_change(regular_value, copy_value),
                ]
            )

    output_workbook = Workbook()
    summary_sheet = output_workbook.active
    summary_sheet.title = "Summary"

    summary_rows = [
        ["Comparison item", "Result"],
        ["Regular workbook", str(regular_path)],
        ["Copy workbook", str(copy_path)],
        ["Ignored data", "Entire first column in both files"],
        ["Regular sheet", regular_sheet.title],
        ["Copy sheet", copy_sheet.title],
        ["Regular header row", regular_header_row],
        ["Copy header row", copy_header_row],
        ["Matched unique books", len(common_book_ids)],
        ["Changed cells", len(differences)],
        ["Books only in Regular", len(regular_only_book_ids)],
        ["Books only in Copy", len(copy_only_book_ids)],
        ["Columns only in Regular", len(regular_only_headers)],
        ["Columns only in Copy", len(copy_only_headers)],
        ["Regular duplicate Book IDs", len(regular_duplicates)],
        ["Copy duplicate Book IDs", len(copy_duplicates)],
        ["Regular nonblank rows without Book ID", len(regular_unkeyed)],
        ["Copy nonblank rows without Book ID", len(copy_unkeyed)],
        [
            "Important formula note",
            (
                "Comparison uses Excel's last saved formula results. "
                "Save both files in Excel before running."
            ),
        ],
    ]

    for row in summary_rows:
        summary_sheet.append(row)

    for cell in summary_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.font = Font(color="FFFFFF", bold=True)

    summary_sheet.freeze_panes = "A2"
    autosize(summary_sheet)

    differences_sheet = output_workbook.create_sheet("Cell Differences")
    make_header(
        differences_sheet,
        [
            "Book ID",
            "Title",
            "Author",
            "Column",
            "Regular Excel row",
            "Copy Excel row",
            "Regular value",
            "Copy value",
            "Change type",
        ],
    )

    for difference in differences:
        differences_sheet.append(difference)

    autosize(differences_sheet)

    write_full_rows_sheet(
        output_workbook,
        "Only in Regular",
        regular_sheet,
        regular_only_book_ids,
        regular_rows,
        regular_headers,
        regular_display_headers,
    )

    write_full_rows_sheet(
        output_workbook,
        "Only in Copy",
        copy_sheet,
        copy_only_book_ids,
        copy_rows,
        copy_headers,
        copy_display_headers,
    )

    column_sheet = output_workbook.create_sheet("Column Differences")
    make_header(column_sheet, ["Location", "Column"])

    for header in sorted(
        regular_only_headers,
        key=lambda item: regular_headers[item],
    ):
        column_sheet.append(
            ["Only in Regular", regular_display_headers[header]]
        )

    for header in sorted(
        copy_only_headers,
        key=lambda item: copy_headers[item],
    ):
        column_sheet.append(
            ["Only in Copy", copy_display_headers[header]]
        )

    autosize(column_sheet)

    duplicate_sheet = output_workbook.create_sheet("Duplicate Book IDs")
    make_header(
        duplicate_sheet,
        ["Workbook", "Book ID", "Excel rows"],
    )

    for book_id, rows in sorted(regular_duplicates.items()):
        duplicate_sheet.append(
            ["Regular", book_id, ", ".join(map(str, rows))]
        )

    for book_id, rows in sorted(copy_duplicates.items()):
        duplicate_sheet.append(
            ["Copy", book_id, ", ".join(map(str, rows))]
        )

    autosize(duplicate_sheet)

    unkeyed_sheet = output_workbook.create_sheet("Rows Without Book ID")
    make_header(
        unkeyed_sheet,
        ["Workbook", "Excel row", "Title", "Author"],
    )

    for label, source_sheet, rows, headers in (
        ("Regular", regular_sheet, regular_unkeyed, regular_headers),
        ("Copy", copy_sheet, copy_unkeyed, copy_headers),
    ):
        for row_number in rows:
            title = get_value(
                source_sheet,
                row_number,
                headers,
                "title",
            )
            first_name = get_value(
                source_sheet,
                row_number,
                headers,
                "first",
            )
            last_name = get_value(
                source_sheet,
                row_number,
                headers,
                "last",
            )
            author = " ".join(
                str(part).strip()
                for part in (first_name, last_name)
                if not is_blank(part)
            )

            unkeyed_sheet.append(
                [
                    label,
                    row_number,
                    display_value(title),
                    author,
                ]
            )

    autosize(unkeyed_sheet)

    settings_sheet = output_workbook.create_sheet("Sheet Settings")
    make_header(
        settings_sheet,
        ["Setting", "Regular", "Copy", "Different?"],
    )

    regular_settings = sheet_settings(regular_sheet)
    copy_settings = sheet_settings(copy_sheet)

    for setting_name in sorted(
        set(regular_settings) | set(copy_settings)
    ):
        regular_setting = regular_settings.get(setting_name, "")
        copy_setting = copy_settings.get(setting_name, "")

        settings_sheet.append(
            [
                setting_name,
                regular_setting,
                copy_setting,
                "YES" if regular_setting != copy_setting else "",
            ]
        )

    autosize(settings_sheet)

    # Make obvious differences easier to spot.
    warning_fill = PatternFill("solid", fgColor="FFF2CC")

    for sheet_name in (
        "Cell Differences",
        "Column Differences",
        "Duplicate Book IDs",
        "Rows Without Book ID",
        "Sheet Settings",
    ):
        sheet = output_workbook[sheet_name]

        for row in sheet.iter_rows(min_row=2):
            if any(not is_blank(cell.value) for cell in row):
                if sheet_name != "Sheet Settings" or row[3].value == "YES":
                    for cell in row:
                        cell.fill = warning_fill

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        output_workbook.save(output_path)
    except PermissionError:
        print(
            f"\nERROR: Could not write the report:\n  {output_path}\n\n"
            "Close the existing comparison report if it is open in Excel."
        )
        return 1

    print("\nCOMPARISON COMPLETE")
    print(f"Changed cells:         {len(differences)}")
    print(f"Books only in Regular: {len(regular_only_book_ids)}")
    print(f"Books only in Copy:    {len(copy_only_book_ids)}")
    print(f"\nReport created:\n  {output_path}\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())