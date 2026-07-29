"""
Read-only workbook discovery and clerical scanning for Library Clerk.

This module never saves a workbook.

openpyxl is used only to read workbook values and structure. Any future
workbook-writing features must use narrowly targeted Excel COM
automation instead.
"""

from __future__ import annotations

import re
import warnings
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import load_workbook

from library_clerk.models import (
    ScanIssue,
    ScanResult,
)


SYSTEM_BOUNDARY_HEADER = (
    "SYSTEM COLUMNS - AUTOMATION ONLY"
)

BOOK_ID_PATTERN = re.compile(
    r"book-[0-9a-f]{8}-"
    r"[0-9a-f]{4}-"
    r"[0-9a-f]{4}-"
    r"[0-9a-f]{4}-"
    r"[0-9a-f]{12}",
    re.IGNORECASE,
)


KNOWN_DATA_HEADERS = (
    "LGBTQ+",
    "ISBN",
    "Year",
    "Pages",
    "Title",
    "Series",
    "First",
    "Last",
    "Genre",
    "Subgenre",
    "Publisher",
    "Origin",
    "Bookcase",
    "Shelf",
    "Position",
)


KNOWN_SYSTEM_HEADERS = (
    "Book ID",
    "Series Sort",
    "Volume Sort",
    "Last Sort",
    "First Sort",
)


REQUIRED_SCAN_COLUMNS = (
    "ISBN",
    "Year",
    "Title",
    "Bookcase",
    "Shelf",
    "Position",
    "Book ID",
)


ProgressCallback = Callable[
    [str],
    None,
]


@dataclass(frozen=True)
class ListViewCandidate:
    """
    One worksheet that structurally resembles List View.
    """

    workbook_path: Path
    sheet_name: str
    headers: tuple[str, ...]
    score: int


def clean(
    value: Any,
) -> str:
    """
    Convert a workbook value to trimmed text.
    """

    if value is None:
        return ""

    return str(
        value
    ).strip()


def normalize_header(
    value: Any,
) -> str:
    """
    Normalize a header for case-insensitive structural matching.
    """

    return re.sub(
        r"[^a-z0-9]+",
        "",
        clean(
            value
        ).casefold(),
    )


def is_missing(
    value: Any,
) -> bool:
    """
    Return True when a cell contains no meaningful clerical value.
    """

    return clean(
        value
    ).casefold() in {
        "",
        "unknown",
        "n/a",
        "na",
        "none",
    }


def read_headers(
    sheet,
) -> list[str]:
    """
    Read and trim the first worksheet row.
    """

    header_row = next(
        sheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        ),
        (),
    )

    headers = [
        clean(
            value
        )
        for value in header_row
    ]

    while (
        headers
        and not headers[-1]
    ):
        headers.pop()

    return headers


def build_header_lookup(
    headers: list[str],
) -> tuple[
    dict[str, int],
    list[str],
]:
    """
    Return normalized header indexes and duplicate header names.

    Indexes are zero-based here and converted to Excel's one-based
    column numbers only when creating a finding.
    """

    header_indexes: dict[
        str,
        int,
    ] = {}

    duplicate_headers: list[str] = []

    for index, header in enumerate(
        headers
    ):
        normalized = normalize_header(
            header
        )

        if not normalized:
            continue

        if normalized in header_indexes:
            duplicate_headers.append(
                header
            )

            continue

        header_indexes[
            normalized
        ] = index

    return (
        header_indexes,
        duplicate_headers,
    )


def get_header_index(
    header_indexes: dict[str, int],
    *header_names: str,
) -> int | None:
    """
    Resolve the first available header alias.
    """

    for header_name in header_names:
        index = header_indexes.get(
            normalize_header(
                header_name
            )
        )

        if index is not None:
            return index

    return None


def get_row_value(
    row_values: tuple[Any, ...],
    index: int | None,
) -> Any:
    """
    Safely read a row value using a zero-based index.
    """

    if index is None:
        return ""

    if index >= len(
        row_values
    ):
        return ""

    return row_values[
        index
    ]


def open_read_only_workbook(
    workbook_path: Path,
):
    """
    Open a workbook without permitting Library Clerk to save it.

    The known unsupported Data Validation extension warning is hidden
    because this application never saves through openpyxl.
    """

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=(
                ".*Data Validation extension "
                "is not supported.*"
            ),
        )

        return load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )


def candidate_score(
    sheet_name: str,
    header_indexes: dict[str, int],
) -> int:
    """
    Score how closely a worksheet resembles List View.
    """

    score = 0

    if sheet_name.casefold() == "list view":
        score += 25

    for header_name in (
        *KNOWN_DATA_HEADERS,
        SYSTEM_BOUNDARY_HEADER,
        *KNOWN_SYSTEM_HEADERS,
    ):
        if (
            normalize_header(
                header_name
            )
            in header_indexes
        ):
            score += 2

    return score


def discover_list_view_candidates(
    workbook_paths: list[Path],
    result: ScanResult,
    progress_callback: ProgressCallback | None,
) -> list[ListViewCandidate]:
    """
    Inspect workbook headers and return possible List View sheets.
    """

    candidates: list[
        ListViewCandidate
    ] = []

    required_headers = {
        normalize_header(
            "Title"
        ),
        normalize_header(
            SYSTEM_BOUNDARY_HEADER
        ),
        normalize_header(
            "Book ID"
        ),
    }

    for workbook_number, workbook_path in enumerate(
        workbook_paths,
        start=1,
    ):
        if progress_callback is not None:
            progress_callback(
                (
                    f"Inspecting workbook "
                    f"{workbook_number} of "
                    f"{len(workbook_paths)}: "
                    f"{workbook_path.name}"
                )
            )

        workbook = None

        try:
            workbook = open_read_only_workbook(
                workbook_path
            )

            for sheet in workbook.worksheets:
                headers = read_headers(
                    sheet
                )

                if not headers:
                    continue

                (
                    header_indexes,
                    _,
                ) = build_header_lookup(
                    headers
                )

                if not required_headers.issubset(
                    header_indexes
                ):
                    continue

                candidates.append(
                    ListViewCandidate(
                        workbook_path=workbook_path,
                        sheet_name=sheet.title,
                        headers=tuple(
                            headers
                        ),
                        score=candidate_score(
                            sheet.title,
                            header_indexes,
                        ),
                    )
                )

        except Exception as error:
            result.warnings.append(
                (
                    f"Could not inspect "
                    f"{workbook_path.name}: "
                    f"{type(error).__name__}: "
                    f"{error}"
                )
            )

        finally:
            if workbook is not None:
                workbook.close()

    return candidates


def select_list_view_candidate(
    candidates: list[ListViewCandidate],
) -> ListViewCandidate | None:
    """
    Select one unambiguous List View candidate.
    """

    if not candidates:
        return None

    sorted_candidates = sorted(
        candidates,
        key=lambda candidate: (
            candidate.score,
            candidate.workbook_path.name.casefold(),
            candidate.sheet_name.casefold(),
        ),
        reverse=True,
    )

    best_candidate = sorted_candidates[
        0
    ]

    equally_ranked = [
        candidate
        for candidate in sorted_candidates
        if candidate.score
        == best_candidate.score
    ]

    if len(
        equally_ranked
    ) > 1:
        return None

    return best_candidate


def make_issue_identity(
    book_id: str,
    row_number: int,
) -> str:
    """
    Return the most stable available identity for an issue.
    """

    normalized_book_id = clean(
        book_id
    ).casefold()

    if BOOK_ID_PATTERN.fullmatch(
        normalized_book_id
    ):
        return normalized_book_id

    return f"row-{row_number}"


def scan_workbooks(
    workbook_paths: list[Path],
    progress_callback: ProgressCallback | None = None,
) -> ScanResult:
    """
    Scan workbook files and return a complete clerical result.
    """

    started_at = perf_counter()

    result = ScanResult(
        workbook_count=len(
            workbook_paths
        )
    )

    try:
        candidates = (
            discover_list_view_candidates(
                workbook_paths,
                result,
                progress_callback,
            )
        )

        candidate = (
            select_list_view_candidate(
                candidates
            )
        )

        if candidate is None:
            if not candidates:
                message = (
                    "No worksheet could be identified as "
                    "List View by its required structure."
                )
            else:
                candidate_names = ", ".join(
                    (
                        f"{item.workbook_path.name}"
                        f" → {item.sheet_name}"
                    )
                    for item in candidates
                )

                message = (
                    "Multiple equally likely List View "
                    "worksheets were found: "
                    f"{candidate_names}"
                )

            result.issues.append(
                ScanIssue(
                    issue_id=(
                        "system:list-view-discovery"
                    ),
                    area="System integrity",
                    category=(
                        "List View discovery"
                    ),
                    message=message,
                    workbook_path=None,
                    sheet_name="",
                )
            )

            return result

        result.list_view_workbook = (
            candidate.workbook_path
        )

        result.list_view_sheet = (
            candidate.sheet_name
        )

        if progress_callback is not None:
            progress_callback(
                (
                    "Scanning List View: "
                    f"{candidate.workbook_path.name}"
                    f" → {candidate.sheet_name}"
                )
            )

        workbook = None

        try:
            workbook = open_read_only_workbook(
                candidate.workbook_path
            )

            sheet = workbook[
                candidate.sheet_name
            ]

            headers = read_headers(
                sheet
            )

            result.discovered_headers = list(
                headers
            )

            (
                header_indexes,
                duplicate_headers,
            ) = build_header_lookup(
                headers
            )

            for duplicate_header in duplicate_headers:
                result.issues.append(
                    ScanIssue(
                        issue_id=(
                            "system:duplicate-header:"
                            f"{normalize_header(duplicate_header)}"
                        ),
                        area="System integrity",
                        category="Duplicate header",
                        message=(
                            "List View contains more than "
                            "one column with the header "
                            f"'{duplicate_header}'."
                        ),
                        workbook_path=(
                            candidate.workbook_path
                        ),
                        sheet_name=(
                            candidate.sheet_name
                        ),
                        row_number=1,
                        column_name=(
                            duplicate_header
                        ),
                    )
                )

            boundary_index = get_header_index(
                header_indexes,
                SYSTEM_BOUNDARY_HEADER,
            )

            book_id_index = get_header_index(
                header_indexes,
                "Book ID",
            )

            title_index = get_header_index(
                header_indexes,
                "Title",
            )

            first_index = get_header_index(
                header_indexes,
                "First",
            )

            last_index = get_header_index(
                header_indexes,
                "Last",
            )

            isbn_index = get_header_index(
                header_indexes,
                "ISBN",
                "ISBN-13",
            )

            year_index = get_header_index(
                header_indexes,
                "Year",
                "Publication Year",
            )

            bookcase_index = get_header_index(
                header_indexes,
                "Bookcase",
            )

            shelf_index = get_header_index(
                header_indexes,
                "Shelf",
            )

            position_index = get_header_index(
                header_indexes,
                "Position",
                "Shelf Position",
            )

            if (
                boundary_index is not None
                and book_id_index is not None
                and book_id_index
                != boundary_index + 1
            ):
                result.issues.append(
                    ScanIssue(
                        issue_id=(
                            "system:book-id-boundary"
                        ),
                        area="System integrity",
                        category=(
                            "Automation boundary"
                        ),
                        message=(
                            "Book ID must appear "
                            "immediately after "
                            f"'{SYSTEM_BOUNDARY_HEADER}'."
                        ),
                        workbook_path=(
                            candidate.workbook_path
                        ),
                        sheet_name=(
                            candidate.sheet_name
                        ),
                        row_number=1,
                        column_name="Book ID",
                        column_number=(
                            None
                            if book_id_index is None
                            else book_id_index + 1
                        ),
                    )
                )

            for required_header in (
                REQUIRED_SCAN_COLUMNS
            ):
                required_index = get_header_index(
                    header_indexes,
                    required_header,
                )

                if required_index is not None:
                    continue

                result.issues.append(
                    ScanIssue(
                        issue_id=(
                            "system:missing-column:"
                            f"{normalize_header(required_header)}"
                        ),
                        area="System integrity",
                        category="Missing column",
                        message=(
                            "List View is missing the "
                            f"'{required_header}' column."
                        ),
                        workbook_path=(
                            candidate.workbook_path
                        ),
                        sheet_name=(
                            candidate.sheet_name
                        ),
                        row_number=1,
                        column_name=(
                            required_header
                        ),
                    )
                )

            if boundary_index is not None:
                known_data = {
                    normalize_header(
                        header
                    )
                    for header in KNOWN_DATA_HEADERS
                }

                known_system = {
                    normalize_header(
                        header
                    )
                    for header in KNOWN_SYSTEM_HEADERS
                }

                for index, header in enumerate(
                    headers
                ):
                    normalized = normalize_header(
                        header
                    )

                    if not normalized:
                        continue

                    if index < boundary_index:
                        if normalized not in known_data:
                            result.additional_data_headers.append(
                                header
                            )

                    elif index > boundary_index:
                        if normalized not in known_system:
                            result.additional_system_headers.append(
                                header
                            )

                for data_header in KNOWN_DATA_HEADERS:
                    data_index = get_header_index(
                        header_indexes,
                        data_header,
                    )

                    if (
                        data_index is not None
                        and data_index
                        > boundary_index
                    ):
                        result.issues.append(
                            ScanIssue(
                                issue_id=(
                                    "system:data-after-boundary:"
                                    f"{normalize_header(data_header)}"
                                ),
                                area="System integrity",
                                category=(
                                    "Automation boundary"
                                ),
                                message=(
                                    f"Clerical data column "
                                    f"'{data_header}' appears "
                                    "after the automation "
                                    "boundary."
                                ),
                                workbook_path=(
                                    candidate.workbook_path
                                ),
                                sheet_name=(
                                    candidate.sheet_name
                                ),
                                row_number=1,
                                column_name=data_header,
                                column_number=(
                                    data_index + 1
                                ),
                            )
                        )

                for system_header in (
                    KNOWN_SYSTEM_HEADERS
                ):
                    system_index = get_header_index(
                        header_indexes,
                        system_header,
                    )

                    if system_index is None:
                        result.issues.append(
                            ScanIssue(
                                issue_id=(
                                    "system:missing-helper:"
                                    f"{normalize_header(system_header)}"
                                ),
                                area="System integrity",
                                category=(
                                    "Missing system column"
                                ),
                                message=(
                                    "List View is missing "
                                    "the system/helper column "
                                    f"'{system_header}'."
                                ),
                                workbook_path=(
                                    candidate.workbook_path
                                ),
                                sheet_name=(
                                    candidate.sheet_name
                                ),
                                row_number=1,
                                column_name=system_header,
                            )
                        )

                    elif (
                        system_index
                        < boundary_index
                    ):
                        result.issues.append(
                            ScanIssue(
                                issue_id=(
                                    "system:helper-before-boundary:"
                                    f"{normalize_header(system_header)}"
                                ),
                                area="System integrity",
                                category=(
                                    "Automation boundary"
                                ),
                                message=(
                                    f"System/helper column "
                                    f"'{system_header}' appears "
                                    "before the automation "
                                    "boundary."
                                ),
                                workbook_path=(
                                    candidate.workbook_path
                                ),
                                sheet_name=(
                                    candidate.sheet_name
                                ),
                                row_number=1,
                                column_name=system_header,
                                column_number=(
                                    system_index + 1
                                ),
                            )
                        )

            first_book_id_rows: dict[
                str,
                int,
            ] = {}

            for row_number, row_values in enumerate(
                sheet.iter_rows(
                    min_row=2,
                    values_only=True,
                ),
                start=2,
            ):
                title = clean(
                    get_row_value(
                        row_values,
                        title_index,
                    )
                )

                book_id = clean(
                    get_row_value(
                        row_values,
                        book_id_index,
                    )
                )

                if not title:
                    if book_id:
                        result.issues.append(
                            ScanIssue(
                                issue_id=(
                                    "system:orphan-book-id:"
                                    f"{row_number}"
                                ),
                                area="System integrity",
                                category=(
                                    "Book ID integrity"
                                ),
                                message=(
                                    "A Book ID is attached "
                                    "to a blank-title row."
                                ),
                                workbook_path=(
                                    candidate.workbook_path
                                ),
                                sheet_name=(
                                    candidate.sheet_name
                                ),
                                row_number=row_number,
                                column_name="Book ID",
                                column_number=(
                                    None
                                    if book_id_index is None
                                    else book_id_index + 1
                                ),
                                book_id=book_id,
                                title="(blank title)",
                            )
                        )

                    continue

                result.books_scanned += 1

                first_name = clean(
                    get_row_value(
                        row_values,
                        first_index,
                    )
                )

                last_name = clean(
                    get_row_value(
                        row_values,
                        last_index,
                    )
                )

                author = " ".join(
                    part
                    for part in (
                        first_name,
                        last_name,
                    )
                    if part
                )

                identity = make_issue_identity(
                    book_id,
                    row_number,
                )

                normalized_book_id = (
                    book_id.casefold()
                )

                book_id_is_valid = (
                    BOOK_ID_PATTERN.fullmatch(
                        normalized_book_id
                    )
                    is not None
                )

                if is_missing(
                    book_id
                ):
                    result.issues.append(
                        ScanIssue(
                            issue_id=(
                                f"row-{row_number}:"
                                "missing-book-id"
                            ),
                            area="System integrity",
                            category=(
                                "Book ID integrity"
                            ),
                            message=(
                                "This titled row is "
                                "missing its permanent "
                                "Book ID."
                            ),
                            workbook_path=(
                                candidate.workbook_path
                            ),
                            sheet_name=(
                                candidate.sheet_name
                            ),
                            row_number=row_number,
                            column_name="Book ID",
                            column_number=(
                                None
                                if book_id_index is None
                                else book_id_index + 1
                            ),
                            title=title,
                            author=author,
                        )
                    )

                elif not book_id_is_valid:
                    result.issues.append(
                        ScanIssue(
                            issue_id=(
                                f"row-{row_number}:"
                                "malformed-book-id"
                            ),
                            area="System integrity",
                            category=(
                                "Book ID integrity"
                            ),
                            message=(
                                "This row contains a "
                                "malformed Book ID."
                            ),
                            workbook_path=(
                                candidate.workbook_path
                            ),
                            sheet_name=(
                                candidate.sheet_name
                            ),
                            row_number=row_number,
                            column_name="Book ID",
                            column_number=(
                                None
                                if book_id_index is None
                                else book_id_index + 1
                            ),
                            book_id=book_id,
                            title=title,
                            author=author,
                            details=book_id,
                        )
                    )

                else:
                    first_row = (
                        first_book_id_rows.get(
                            normalized_book_id
                        )
                    )

                    if first_row is None:
                        first_book_id_rows[
                            normalized_book_id
                        ] = row_number

                    else:
                        result.issues.append(
                            ScanIssue(
                                issue_id=(
                                    f"{normalized_book_id}:"
                                    f"duplicate:{row_number}"
                                ),
                                area="System integrity",
                                category=(
                                    "Book ID integrity"
                                ),
                                message=(
                                    "This Book ID is "
                                    "duplicated. Its first "
                                    "use is on row "
                                    f"{first_row}."
                                ),
                                workbook_path=(
                                    candidate.workbook_path
                                ),
                                sheet_name=(
                                    candidate.sheet_name
                                ),
                                row_number=row_number,
                                column_name="Book ID",
                                column_number=(
                                    None
                                    if book_id_index is None
                                    else book_id_index + 1
                                ),
                                book_id=book_id,
                                title=title,
                                author=author,
                            )
                        )

                clerical_checks = (
                    (
                        "ISBN",
                        isbn_index,
                        "missing-isbn",
                    ),
                    (
                        "Year",
                        year_index,
                        "missing-year",
                    ),
                    (
                        "Bookcase",
                        bookcase_index,
                        "missing-bookcase",
                    ),
                    (
                        "Shelf",
                        shelf_index,
                        "missing-shelf",
                    ),
                )

                for (
                    column_name,
                    column_index,
                    issue_code,
                ) in clerical_checks:
                    if column_index is None:
                        continue

                    value = get_row_value(
                        row_values,
                        column_index,
                    )

                    if not is_missing(
                        value
                    ):
                        continue

                    result.issues.append(
                        ScanIssue(
                            issue_id=(
                                f"{identity}:"
                                f"{issue_code}"
                            ),
                            issue_code=issue_code,
                            area="Clerical",
                            category=(
                                f"Missing {column_name}"
                            ),
                            message=(
                                f"This book is missing "
                                f"{column_name}."
                            ),
                            workbook_path=(
                                candidate.workbook_path
                            ),
                            sheet_name=(
                                candidate.sheet_name
                            ),
                            row_number=row_number,
                            column_name=column_name,
                            column_number=(
                                column_index + 1
                            ),
                            book_id=book_id,
                            title=title,
                            author=author,
                        )
                    )

                bookcase = clean(
                    get_row_value(
                        row_values,
                        bookcase_index,
                    )
                )

                position = get_row_value(
                    row_values,
                    position_index,
                )

                if (
                    position_index is not None
                    and bookcase.casefold()
                    == "office"
                    and is_missing(
                        position
                    )
                ):
                    result.issues.append(
                        ScanIssue(
                            issue_id=(
                                f"{identity}:"
                                "missing-office-position"
                            ),
                            issue_code=(
                                "missing-office-position"
                            ),
                            area="Clerical",
                            category=(
                                "Missing Office Position"
                            ),
                            message=(
                                "Office books require a "
                                "shelf Position."
                            ),
                            workbook_path=(
                                candidate.workbook_path
                            ),
                            sheet_name=(
                                candidate.sheet_name
                            ),
                            row_number=row_number,
                            column_name="Position",
                            column_number=(
                                position_index + 1
                            ),
                            book_id=book_id,
                            title=title,
                            author=author,
                        )
                    )

        except Exception as error:
            result.issues.append(
                ScanIssue(
                    issue_id=(
                        "system:list-view-read-error"
                    ),
                    area="System integrity",
                    category="Workbook read error",
                    message=(
                        "Library Clerk could not finish "
                        "reading List View."
                    ),
                    workbook_path=(
                        candidate.workbook_path
                    ),
                    sheet_name=(
                        candidate.sheet_name
                    ),
                    details=(
                        f"{type(error).__name__}: "
                        f"{error}"
                    ),
                )
            )

        finally:
            if workbook is not None:
                workbook.close()

        return result

    finally:
        result.duration_seconds = (
            perf_counter()
            - started_at
        )